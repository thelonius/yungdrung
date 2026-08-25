#!/usr/bin/env python3
"""Движок шагов Yungdrung.

Единственный, кто пишет шаги и даты в вольт. Без LLM: всё, что здесь считается —
выборки по датам и статусам. На выходе JSON, чтобы поверх можно было повесить
любой интерфейс.

Вывод всегда в UTF-8, независимо от кодировки консоли. Кто вызывает движок из
кода и разбирает JSON — обязан читать его как UTF-8 явно: на Windows кодировка
локали другая, и русский текст молча превратится в «Ð“Ñ€Ð°Ð½Ñ‚».
Для subprocess это `encoding="utf-8"`.

  python3 engine.py next                        что требует внимания сегодня
  python3 engine.py done <задача> <шаг>         шаг сделан
  python3 engine.py notdone <задача> <шаг> --reason "не дозвонился"
  python3 engine.py defer <задача> <шаг> --to 2026-08-20 --reason "..."
  python3 engine.py list                        все задачи со статусами
  python3 engine.py show <задача>               одна задача целиком
  python3 engine.py export --to выгрузка.xlsx   весь вольт в Excel

Задача указывается частью имени файла: "грант" найдёт «Заявка на грант ФПГ».
"""
import argparse
import json
import os
import re
import sys
import tempfile
import time
from datetime import date, datetime, time as dtime, timedelta
from pathlib import Path
from types import SimpleNamespace

# Консоль Windows по умолчанию не в UTF-8 (обычно cp866), а мы печатаем русский
# текст и типографику: «кавычки», тире, многоточие. Без этого «нет задачи по
# «грант»» падает UnicodeEncodeError вместо внятного сообщения — причём на
# машине заказчика, который такое не починит. Делаем до первого вывода.
for stream in (sys.stdout, sys.stderr):
    try:
        stream.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass  # поток подменён или не текстовый — печатать всё равно нечем испортить

try:
    import yaml
except ImportError:
    sys.exit("нужен pyyaml: pip install pyyaml")

import attachments
import backup
import kb
import recurrence as rec
import settings as cfg
import store
import templates as tpl
import worktime

SCHEMA = 1
VAULT = Path(os.environ.get("YUNGDRUNG_VAULT", Path(__file__).resolve().parent))
KB_DIR = VAULT / "База"


def db_path():
    """Путь к базе. Функция, а не константа — по той же причине, что и
    `get_reasons()` ниже: `VAULT` в тестах подменяют через monkeypatch уже
    после импорта, и захардкоженный путь эту подмену не увидит."""
    return VAULT / "вольт.db"


def get_store():
    return store.Store(db_path())

OPEN = "pending"
DONE = "done"
SKIPPED = "skipped"
FAILED = "failed"

def get_reasons():
    """Справочник причин, раздел 5.4 ТЗ. Причина обязательна при «не сделано»,
    переносе и провале: без неё счётчик переносов показывает, что шаг буксует,
    но не показывает, обо что.

    Список редактируется в настройках (settings.py) — здесь только чтение.
    Функция, а не константа: захардкоженный список нельзя переименовать или
    заархивировать из интерфейса, а settings.py уже это умеет. `VAULT` берём
    именно из engine (не из settings.cfg.VAULT — тот вычислен независимо при
    импорте и не отследит подмену в тестах через monkeypatch).
    """
    return cfg.active_reason_names(cfg.settings_path(VAULT))

# В вольт статус пишется по-русски: эти файлы читает заказчик, а не только движок.
# В JSON наружу уходят английские ключи — там интерфейс для бота.
STATUS_RU = {
    "overdue": "просрочена",
    "due": "сегодня",
    "waiting": "ждёт",
    "no_date": "без даты",
    "done": "закрыта",
    "empty": "нет шагов",
    "cancelled": "отменена",
}

# Статусы и события шага в файле лежат по-английски — их читает движок. В Excel
# уходит перевод: тот файл открывает заказчик, и «not_done» ему ни о чём не говорит.
STEP_STATUS_RU = {
    OPEN: "ждёт",
    DONE: "сделан",
    SKIPPED: "снят",
    FAILED: "провален",
}

EVENT_RU = {
    "done": "сделан",
    "failed": "провален",
    "not_done": "не сделан",
    "defer": "перенесён",
    "skipped": "снят",
    "reopened": "переоткрыт",
}


# --- чтение и запись -------------------------------------------------------

def parse_file(path):
    """Frontmatter + тело. Тело сохраняем как есть: его пишет заказчик, не мы."""
    raw = path.read_text(encoding="utf-8")
    if not raw.startswith("---"):
        raise ValueError(f"{path.name}: нет frontmatter")
    _, fm, body = raw.split("---", 2)
    return yaml.safe_load(fm) or {}, body.lstrip("\n")


class PlainDumper(yaml.SafeDumper):
    """Без якорей и ссылок: одна и та же дата в нескольких полях — обычное дело,
    а `&id001`/`*id001` Obsidian не разбирает и файл для него ломается."""

    def ignore_aliases(self, data):
        return True


def write_file(path, meta, body):
    """Атомарно: Obsidian держит файлы открытыми и кэширует, дописывать на месте нельзя.

    На Windows os.replace может ненадолго упасть с PermissionError, если файл в
    этот момент держит антивирус, OneDrive-индексатор или сам Obsidian — на маке
    так почти не бывает. Несколько коротких повторов дешевле, чем терять запись.
    """
    fm = yaml.dump(meta, Dumper=PlainDumper, allow_unicode=True, sort_keys=False,
                   default_flow_style=False)
    content = f"---\n{fm}---\n\n{body}"
    fd, tmp = tempfile.mkstemp(dir=str(path.parent), suffix=".tmp")
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            f.write(content)
        for attempt in range(5):
            try:
                os.replace(tmp, path)
                break
            except PermissionError:
                if attempt == 4:
                    raise
                time.sleep(0.2 * (attempt + 1))
    except BaseException:
        Path(tmp).unlink(missing_ok=True)
        raise


# Задачи теперь читает store.py — SQLite не оставляет полуразобранных строк
# так, как правленный руками YAML оставлял полуразобранные файлы. Список остаётся
# (всегда пустой) ради стабильности формы ответа: `cmd_feed`/`cmd_backlog`/
# `cmd_next`/`cmd_list`/`cmd_refresh` отдают "broken" по контракту, и снимать
# ключ без отдельного решения — не эта задача.
BROKEN = []


def load_tasks():
    return get_store().load_tasks()


def find_task(fragment):
    frag = fragment.lower()
    hits = [t for t in load_tasks() if frag in t["path"].stem.lower()]
    if not hits:
        sys.exit(f"нет задачи по «{fragment}»")
    if len(hits) > 1:
        sys.exit("подходит несколько: " + ", ".join(t["path"].stem for t in hits))
    return hits[0]


# Записи базы знаний, которые не разобрались при последнем чтении. Та же логика,
# что у BROKEN: опечатка в заметке заказчика не должна тихо выкидывать запись
# из автораспознавания без единого сигнала об этом.
KB_BROKEN = []


def load_kb_entries():
    """Записи базы знаний для kb.build_index — раздел 5.7 ТЗ.

    Источник — база, если в неё уже переехали, иначе `База/*.md`. Порядок
    именно такой: этап (b) переносит записи один раз, и после переноса markdown
    остаётся на диске как был (мы его не удаляем — это данные заказчика), но
    источником правды перестаёт быть. Читать оба разом нельзя: одна и та же
    запись дала бы два совпадения в тексте.
    """
    KB_BROKEN.clear()
    из_базы = get_store().load_kb_notes()
    if из_базы:
        return [{"id": з["id"], "title": з["title"], "aliases": з["aliases"]}
                for з in из_базы]
    if not KB_DIR.is_dir():
        return []
    out = []
    for path in sorted(KB_DIR.glob("*.md")):
        try:
            meta, _ = parse_file(path)
        except Exception as e:
            reason = " ".join(str(e).split())[:200]
            KB_BROKEN.append({"file": path.name, "error": reason})
            continue
        if meta.get("type") != "note":
            continue
        title = (meta.get("title") or "").strip()
        if not title:
            KB_BROKEN.append({"file": path.name, "error": "нет названия"})
            continue
        out.append({
            "id": path.stem,
            "title": title,
            "aliases": meta.get("aliases") or meta.get("synonyms") or [],
        })
    return out


# --- вычисления ------------------------------------------------------------

def as_date(value):
    """Дата контроля может быть с временем или без: без времени — весь день."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return datetime.fromisoformat(str(value).strip()).date()


def parse_stored_control(text):
    """Строка control_date (из `parse_date_input` или уже из хранилища) →
    `date` или `datetime`, время сохраняется, если было.

    Отдельно от `as_date` намеренно: тот всегда режет время, а перенос шага
    («перенести на 15:00», пресет «через час») обязан его пронести. Раньше
    `cmd_defer`/`cmd_notdone` звали `date.fromisoformat(args.to)` напрямую —
    на строке без времени это работало, а на «2026-08-24 15:00» падало с
    ValueError, и «перенести на конкретный час» было в принципе недостижимо.

    Порядок проверки важен: `date.fromisoformat` на строке с временем сам
    бросает ValueError, и код проваливается в `datetime.fromisoformat`,
    который с Python 3.11 принимает и пробел, и «T» как разделитель.
    """
    s = str(text).strip()
    try:
        return date.fromisoformat(s)
    except ValueError:
        return datetime.fromisoformat(s)


def steps_of(task):
    return task["meta"].get("steps") or []


def is_closed(step):
    """Шаг закрыт, только если статус — один из двух известных нам.

    Всё незнакомое считаем открытым, а не закрытым. Вольт правится руками, и
    заказчик вполне может напечатать в Obsidian `status: сделан` вместо `done`.
    Раньше это роняло `list`/`next`/`refresh`/`export` целиком с traceback:
    «закрыт» и «открыт» проверялись двумя разными правилами, и на незнакомом
    статусе они расходились — задача не считалась закрытой, но и текущего шага
    в ней не находилось.
    """
    return step.get("status", OPEN) in (DONE, SKIPPED, FAILED)


def is_group(step):
    """Группа подшагов. У группы есть режим — "par" (порядок не важен) или
    "seq" (подшаги по очереди), — а дат, статуса и журнала нет: закрытие
    вычисляется из детей. Обычный шаг режима не имеет."""
    return bool(step.get("mode"))


def _children_map(steps):
    """id родителя → его дети в порядке хранения. Верхний уровень — под ключом
    None. Плоский список из БД идёт в глубину-первом порядке, поэтому дети
    каждого родителя здесь оказываются в своём относительном порядке."""
    m = {}
    for s in steps:
        m.setdefault(s.get("parent"), []).append(s)
    return m


def _closure(steps):
    """(закрыт?, карта детей) для дерева шагов. Лист закрыт по статусу, группа —
    когда закрыты все дети. Группа без детей считается закрытой: валидация
    такие не пропускает, а на битых данных «закрыта» безопаснее вечно
    открытой — не всплывает в ленте."""
    m = _children_map(steps)

    def закрыт(s):
        if is_group(s):
            return all(закрыт(c) for c in m.get(s["id"], []))
        return is_closed(s)

    return закрыт, m


def leaves_of(task):
    return [s for s in steps_of(task) if not is_group(s)]


def current_steps(task):
    """Активные листья — то, по чему сейчас идёт работа. В последовательной
    цепочке это листья первого незакрытого элемента; параллельная группа
    отдаёт активные листья всех своих незакрытых детей разом."""
    steps = steps_of(task)
    закрыт, m = _closure(steps)

    def раскрыть(s):
        if not is_group(s):
            return [s]
        дети = [c for c in m.get(s["id"], []) if not закрыт(c)]
        if s["mode"] == "par":
            return [лист for c in дети for лист in раскрыть(c)]
        return раскрыть(дети[0]) if дети else []

    for s in m.get(None, []):
        if not закрыт(s):
            return раскрыть(s)
    return []


def current_step(task):
    """Первый активный лист — для сводки и мест, где нужен один шаг."""
    активные = current_steps(task)
    return активные[0] if активные else None


def task_status(task, today):
    # Отмена — состояние задачи целиком, не выводится из шагов и стоит впереди
    # любого другого правила: отменённая задача не должна всплывать просроченной
    # только потому, что в ней остался незакрытый шаг.
    if task["meta"].get("cancelled"):
        return "cancelled"
    steps = steps_of(task)
    if not steps:
        return "empty"
    активные = current_steps(task)
    if not активные:
        return "done"
    # Активных листьев может быть несколько (параллельная группа) — задача
    # получает худшее из их состояний: просрочка перекрывает «сегодня», та —
    # отсутствие даты, та — ожидание. Один лист даёт прежнее поведение.
    состояния = set()
    for step in активные:
        due = as_date(step.get("control_date"))
        if due is None:
            состояния.add("no_date")
        elif due < today:
            состояния.add("overdue")
        elif due == today:
            состояния.add("due")
        else:
            состояния.add("waiting")
    for худшее in ("overdue", "due", "no_date"):
        if худшее in состояния:
            return худшее
    return "waiting"


def stall_count(step):
    """Сколько раз шаг не сделали. Отличает «ещё не дошли руки» от «буксует».

    Массовый перенос (`mass_defer`) считается наравне с «не сделан» — R20 ТЗ
    прямо требует, чтобы он увеличивал счётчик. Обычный одиночный `defer` сюда
    не идёт: там дату для конкретного шага выбирают осознанно, это не то же
    самое, что «опять не собрались» (см. test_defer_does_not_count_as_stalling).
    """
    return sum(1 for e in step.get("log") or [] if e.get("event") in ("not_done", "mass_defer"))


def step_view(task, step, today):
    due = as_date(step.get("control_date"))
    return {
        "task": task["path"].stem,
        "step": step.get("id"),
        "title": step.get("title"),
        "control_date": str(step.get("control_date")) if step.get("control_date") else None,
        "overdue_days": (today - due).days if due and due < today else 0,
        "stalled": stall_count(step),
        "last_reason": next(
            (e.get("reason") for e in reversed(step.get("log") or []) if e.get("reason")), None
        ),
    }


# --- запись ----------------------------------------------------------------

def log_event(step, event, today, **fields):
    step.setdefault("log", [])
    entry = {"date": today, "event": event}
    entry.update({k: v for k, v in fields.items() if v is not None})
    step["log"].append(entry)


def get_step(task, step_id):
    """Найти шаг для отметки. Группа отметки не принимает — done/defer и
    остальные работают по её подшагам, а закрытие группы вычисляется."""
    for step in steps_of(task):
        if str(step.get("id")) == str(step_id):
            if is_group(step):
                sys.exit(f"шаг {step_id} — группа, отмечаются её подшаги")
            return step
    sys.exit(f"нет шага {step_id} в «{task['path'].stem}»")


STEPS_START = "<!-- шаги: пишет движок, править руками не нужно -->"
STEPS_END = "<!-- /шаги -->"

# Маркеры статуса шага в теле заметки. Галочки-чекбоксы намеренно не используем:
# в Obsidian они кликабельные, заказчик отметил бы шаг мышкой, движок затёр бы
# это при следующей записи — и получилось бы два писателя одного поля.
MARK = {DONE: "✓", SKIPPED: "×", FAILED: "✗", OPEN: "•"}


def render_steps(task, today):
    """Шаги человеческим списком — в тело заметки.

    Зачем: в панели свойств Obsidian массив объектов не рисуется, там видна
    обрезанная JSON-строка. Заказчик кликает по задаче из таблицы Bases и
    упирается в нечитаемое. Здесь он видит список.

    Побочно чинится вторая дыра: `[[ссылка]]` внутри названия шага живёт во
    frontmatter, а его Obsidian ссылками не считает. В теле — считает, поэтому
    ссылки на заметки базы знаний из названий шагов начинают работать.
    """
    lines = [STEPS_START, "**Шаги**", ""]
    for step in steps_of(task):
        status = step.get("status", OPEN)
        tail = []
        due = as_date(step.get("control_date"))
        if status == DONE:
            completed = as_date(step.get("completed_date"))
            tail.append(f"сделан {completed:%d.%m.%Y}" if completed else "сделан")
        elif status == SKIPPED:
            tail.append("снят")
        elif status == FAILED:
            tail.append("не будет сделан")
        elif due:
            tail.append(f"контроль {due:%d.%m.%Y}")
            if due < today:
                tail.append(f"просрочен на {(today - due).days} дн.")
        else:
            tail.append("дата не назначена")

        stalled = stall_count(step)
        if stalled >= 3 and status == OPEN:
            reason = next((e.get("reason") for e in reversed(step.get("log") or [])
                            if e.get("reason")), None)
            tail.append(f"буксует, отметок «не сделан»: {stalled}"
                         + (f" ({reason})" if reason else ""))

        lines.append(f"{MARK.get(status, '•')} **{step.get('id')}.** "
                      f"{step.get('title', '')} — {' · '.join(tail)}")
    lines += ["", STEPS_END]
    return "\n".join(lines)


def put_steps_into_body(body, block):
    """Блок шагов переписываем, всё остальное в теле не трогаем.

    Тело принадлежит заказчику: там его заметки по задаче. Движок владеет только
    участком между маркерами. Если маркеров нет — вставляем блок сверху, текст
    заказчика уезжает под него.
    """
    start = body.find(STEPS_START)
    end = body.find(STEPS_END)
    if start != -1 and end != -1 and end > start:
        tail = body[end + len(STEPS_END):]
        return body[:start] + block + tail
    return block + "\n\n" + body.lstrip("\n") if body.strip() else block + "\n"


def task_summary(task, today):
    """status/current_step/control_date/stalled/progress — сводка верхнего
    уровня, которую `save()` мержит в `task["meta"]` перед возвратом.

    Отдельная функция, а не кусок внутри save(): раньше сводку было видно
    только прочитав только что написанный файл, теперь — сразу после чтения
    из БД её там нет (колонок под неё нет, source of truth в шагах). Тестам и
    любому будущему коду, которому нужен «файл как он раньше выглядел бы»,
    нужен ровно этот пересчёт, а не второе его написание.
    """
    листья = leaves_of(task)
    активные = current_steps(task)
    step = активные[0] if активные else None
    closed = sum(1 for s in листья if is_closed(s))
    # Прогресс считается по листьям: группа — скобка вокруг подшагов, а не
    # отдельная единица работы. Контроль сводки — ближайший из активных.
    контроли = sorted((as_date(s["control_date"]) for s in активные
                       if s.get("control_date")))
    return {
        "status": STATUS_RU[task_status(task, today)],
        "current_step": step.get("title") if step else None,
        "control_date": контроли[0] if контроли else None,
        "stalled": max((stall_count(s) for s in активные), default=0),
        "progress": f"{closed}/{len(листья)}" if листья else None,
    }


def save(task, today):
    """Статус и сводка пересчитываются при каждой записи, руками их никто не ставит.

    Сводка в БД не хранится вообще — колонок под неё нет, это была чистая
    денормализация под таблицу Obsidian Bases. Здесь она по-прежнему мержится в
    `task["meta"]`, потому что вызывающий код (`cmd_update`/`cmd_cancel`/
    `cmd_reopen` и другие) читает `task["meta"]["status"]` сразу после save() —
    источник правды остаётся в шагах, а это просто удобный снимок для JSON.
    """
    meta = task["meta"]
    meta["schema"] = SCHEMA
    meta.update(task_summary(task, today))
    склад = get_store()
    склад.save_task(task, today)
    _index_task(склад, task)


def _index_task(склад, task):
    """Обновить поисковый индекс по одной задаче — сразу после её записи.

    Одна строка, а не пересборка всего: `reindex_search` нужен после переезда и
    для починки, но платить им за каждую отметку шага нельзя. Индексируется
    новое название, поэтому строка сначала удаляется по нему же — при
    переименовании старая осталась бы висеть и находилась по прежнему слову.
    Отсюда `_forget_old_title`: имя до правки знает только вызывающий.

    Сбой индексации не роняет запись: задача уже в базе, и потерять её из-за
    того, что не собрались леммы, было бы хуже, чем разойтись с индексом —
    индекс чинится командой `reindex`, а задача ничем.
    """
    try:
        склад.search_replace(
            "task", task["path"].stem, task["path"].stem,
            (task.get("body") or "").strip()[:200],
            kb.lemmatize_text(_search_text_of_task(task)))
    except Exception:
        pass


# --- команды ---------------------------------------------------------------

def feed_item(task, step, now, work, group=None):
    """Строка ленты. Всё вычислено здесь: морда только показывает.

    Раздел 6.1 ТЗ перечисляет, что видно в строке: название шага, название задачи,
    время контроля, теги, счётчик переносов, если больше нуля.
    """
    control = step.get("control_date")
    показ = worktime.show_at(control, work) if control else None
    return {
        "task": task["path"].stem,
        "step": step.get("id"),
        "title": step.get("title"),
        "group": group,
        "note": step.get("note"),
        "control_at": str(control) if control else None,
        "show_at": показ.isoformat() if показ else None,
        "state": worktime.due_state(control, now, work),
        "postponed": stall_count(step),
        "stalled": stall_count(step) >= 3,
        "tags": task["meta"].get("tags") or [],
        "last_reason": next(
            (e.get("reason") for e in reversed(step.get("log") or []) if e.get("reason")),
            None),
        "actions": ["done", "notdone", "defer", "skip"],
    }


def collect_open(now, work):
    """Открытые шаги всех задач, разложенные по состоянию.

    Одним проходом, потому что лента и завал — это один и тот же набор, просто
    разрезанный по-разному, и считать его дважды значит однажды разойтись.
    """
    лента, завал, ждут = [], [], []
    for task in load_tasks():
        # Отменённая задача уходит из всех трёх наборов целиком. Шаг в ней
        # остаётся открытым — по нему просто больше не работают, — поэтому без
        # явной проверки она каждый день всплывала бы просроченной, хотя
        # `task_status` считает её отменённой первым же правилом. Заодно это
        # снимает напоминания: `remind.py` ходит за списком сюда же.
        if task["meta"].get("cancelled"):
            continue
        по_id = {s["id"]: s for s in steps_of(task)}
        # Параллельная группа даёт несколько активных листьев — и несколько
        # строк ленты: у каждого свой срок, прятать их друг за друга нечестно.
        # Название группы едет в строку контекстом.
        for step in current_steps(task):
            родитель = по_id.get(step.get("parent"))
            item = feed_item(task, step, now, work,
                             group=родитель.get("title") if родитель else None)
            if item["state"] == "overdue":
                завал.append(item)
            elif worktime.in_horizon(step.get("control_date"), now, work):
                лента.append(item)
            else:
                ждут.append(item)
    ключ = lambda i: (i["show_at"] or "9999", i["task"], i["step"] or 0)
    return sorted(лента, key=ключ), sorted(завал, key=ключ), sorted(ждут, key=ключ)


def cmd_feed(args, today):
    """Лента «Что сегодня» — раздел 6.1 ТЗ.

    Просроченное в строки не попадает: по 6.1 оно живёт отдельной плашкой, потому
    что пятнадцать красных строк парализуют экран. Счётчик отдаёт отдельно.
    """
    now = _now(args, today)
    work = _work(args)
    лента, завал, ждут = collect_open(now, work)
    ближайший = ждут[0] if ждут else None
    return {
        "now": now.isoformat(),
        "feed": лента,
        "overdue_count": len(завал),
        "counts": {"overdue": len(завал), "today": len(лента), "waiting": len(ждут)},
        "next_ahead": ближайший,
        "stalled_count": sum(1 for i in лента + завал if i["stalled"]),
        "broken": list(BROKEN),
    }


def cmd_backlog(args, today):
    """Разбор завала — раздел 6.9 ТЗ. Сортировка по умолчанию — сначала самое
    давнее, то есть по `show_at` возрастанием: пункт ТЗ явный и без исключений
    для буксующих. Раньше буксующие элементы всплывали наверх поверх более
    старых просрочек (`(not stalled, show_at)`) — это удобно интуитивно, но
    противоречит явно записанному правилу, и ни один тест этого не стерёг.
    """
    now = _now(args, today)
    work = _work(args)
    _, завал, _ = collect_open(now, work)
    завал.sort(key=lambda i: i["show_at"] or "9999")
    return {"now": now.isoformat(), "backlog": завал, "count": len(завал),
            "broken": list(BROKEN)}


def _now(args, today):
    """Момент, от которого считаем. `--today` задаёт дату, время берём текущее —
    так тесты и утренние прогоны воспроизводимы, а живой запуск точен."""
    заданный = getattr(args, "now", None) if args else None
    if заданный:
        return worktime.as_datetime(заданный)
    сейчас = datetime.now()
    return сейчас if today == сейчас.date() else datetime.combine(today, сейчас.time())


def _work(args):
    """Рабочие часы: настройки из файла — база, аргументы вызова — оверрайд
    поверх них. Раньше файл настроек не читался вовсе, и `Настройки.json` мог
    хранить что угодно — трекер всё равно жил на 09:00–21:00 по умолчанию.
    """
    try:
        сохранённые = cfg.load(cfg.settings_path(VAULT))["notifications"]
    except cfg.SettingsError:
        # Битый файл настроек не должен останавливать ленту и завал — почему
        # он битый, разбирается в настройках-интерфейсе, а не здесь.
        сохранённые = cfg.defaults()["notifications"]

    def выбрать(из_аргумента, ключ):
        return из_аргумента if из_аргумента is not None else сохранённые.get(ключ)

    return worktime.settings(
        start=выбрать(getattr(args, "work_start", None) if args else None, "start"),
        end=выбрать(getattr(args, "work_end", None) if args else None, "end"),
        weekends=выбрать(getattr(args, "weekends", None) if args else None, "weekends"),
    )


def cmd_next(args, today):
    due, stalled = [], []
    for task in load_tasks():
        status = task_status(task, today)
        if status not in ("overdue", "due", "no_date"):
            continue
        for step in current_steps(task):
            view = step_view(task, step, today)
            view["status"] = status
            due.append(view)
            if view["stalled"] >= 3:
                stalled.append(view)
    due.sort(key=lambda v: (-v["overdue_days"], v["task"]))
    return {"today": today.isoformat(), "due": due, "stalled": stalled,
            "broken": list(BROKEN)}


ДНИ_НЕДЕЛИ = {
    "пн": 0, "понедельник": 0, "вт": 1, "вторник": 1, "ср": 2, "среда": 2,
    "чт": 3, "четверг": 3, "пт": 4, "пятница": 4, "сб": 5, "суббота": 5,
    "вс": 6, "воскресенье": 6,
}

ОТНОСИТЕЛЬНЫЕ = {"сегодня": 0, "завтра": 1, "послезавтра": 2}

# Месяцы словом сравниваются по основе, а не полной таблицей падежей: «марта»,
# «марте», «март» — один и тот же месяц, и хвост можно просто отбросить. Приём
# взят из разбора дат в adaptive-astro-scheduler, где он уже отработал.
#
# «Май» вынесен отдельно, и это не педантизм: его основа «ма» короче любой
# другой и своим хвостом съедает «марта» — ровно тот баг, который в исходном
# коде и живёт. Три формы перечислить дешевле, чем сторожить исключение.
МЕСЯЦЫ_ОСНОВЫ = [
    ("январ", 1), ("феврал", 2), ("март", 3), ("апрел", 4), ("июн", 6),
    ("июл", 7), ("август", 8), ("сентябр", 9), ("октябр", 10),
    ("ноябр", 11), ("декабр", 12),
]
МЕСЯЦЫ_МАЙ = {"май": 5, "мая": 5, "мае": 5}


def _месяц(слово):
    """Название месяца в любом падеже → номер, иначе None.

    Хвост после основы ограничен двумя буквами: без этого «мартышка» прошла бы
    за март. Совсем строгой проверки падежа тут не нужно — дальше по разбору
    стоит номер дня, и мусор всё равно не соберётся в дату.
    """
    if слово in МЕСЯЦЫ_МАЙ:
        return МЕСЯЦЫ_МАЙ[слово]
    for основа, номер in МЕСЯЦЫ_ОСНОВЫ:
        if not слово.startswith(основа):
            continue
        хвост = слово[len(основа):]
        if хвост == "" or (len(хвост) <= 2 and хвост.isalpha()):
            return номер
    return None

# Предлоги-связки. У шага одна точка времени, а не диапазон, поэтому «в 9»,
# «до 9» и «к 9» называют один и тот же час: разница есть для человека, для
# движка её нет. Выкидываем их и разбираем то, что осталось — иначе «завтра в
# 9-30» падает на слове «в», хотя и дата, и время в строке написаны.
СВЯЗКИ = {"в", "во", "на", "до", "после", "к", "ко"}

# Время суток словом: «7 вечера» пишут чаще, чем «19:00».
ЧАСТИ_СУТОК = {"утра", "дня", "вечера", "ночи"}

# Названия часов — просто другая запись времени, разворачиваем до разбора.
ИМЕНА_ЧАСОВ = {"полдень": "12:00", "полночь": "00:00"}

# Разговорное время. «Полдесятого» — половина ДЕСЯТОГО часа, то есть 9:30:
# порядковое число называет час, который идёт, а не который прошёл. Ровно та же
# логика у «без пятнадцати десять» (9:45) и «четверть десятого» (9:15) — везде
# считается от следующего часа назад.
#
# Половина суток не угадывается: «полдесятого» — это 9:30, а не 21:30, ровно
# как «в 9» даёт 9:00, а не 21:00. Кому нужен вечер, тот пишет «полдесятого
# вечера», и слово доворачивает час обычным путём (`_время_суток`). Угадывать
# по рабочему дню — значит иногда ставить контроль на двенадцать часов мимо,
# причём молча.
ЧАСЫ_ПОРЯДКОВЫЕ = {
    "первого": 1, "второго": 2, "третьего": 3, "четвертого": 4, "пятого": 5,
    "шестого": 6, "седьмого": 7, "восьмого": 8, "девятого": 9, "десятого": 10,
    "одиннадцатого": 11, "двенадцатого": 12,
}
ЧАСЫ_КОЛИЧЕСТВЕННЫЕ = {
    "час": 1, "часа": 1, "два": 2, "три": 3, "четыре": 4, "пять": 5,
    "шесть": 6, "семь": 7, "восемь": 8, "девять": 9, "десять": 10,
    "одиннадцать": 11, "двенадцать": 12,
}
# Минуты в родительном — то, что стоит после «без». «Четверти» тут же: для
# разбора это просто пятнадцать, написанное словом.
МИНУТЫ_РОДИТЕЛЬНЫЕ = {
    "пяти": 5, "десяти": 10, "четверти": 15, "пятнадцати": 15,
    "двадцати": 20, "двадцати пяти": 25,
}
ПОЛОВИНА_СЛОВОМ = {"пол", "половина", "половине", "половины"}


def _пол_часа(слова, i):
    """«полдесятого», «пол десятого», «пол-десятого», «половине десятого»."""
    слово = слова[i]
    if слово.startswith("пол"):
        # «полдень» и «полночь» до сюда не доходят — их развернул ИМЕНА_ЧАСОВ.
        час = ЧАСЫ_ПОРЯДКОВЫЕ.get(слово[3:].lstrip("-"))
        if час:
            return f"{час - 1:02d}:30", 1
    if слово in ПОЛОВИНА_СЛОВОМ and i + 1 < len(слова):
        час = ЧАСЫ_ПОРЯДКОВЫЕ.get(слова[i + 1])
        if час:
            return f"{час - 1:02d}:30", 2
    return None


def _четверть_часа(слова, i):
    """«четверть десятого» → 9:15. «Без четверти» разбирает `_без_минут`."""
    if слова[i] in ("четверть", "четверти") and i + 1 < len(слова):
        час = ЧАСЫ_ПОРЯДКОВЫЕ.get(слова[i + 1])
        if час:
            return f"{час - 1:02d}:15", 2
    return None


def _без_минут(слова, i):
    """«без пятнадцати десять» → 9:45, «без двадцати пяти шесть» → 5:35.

    Минуты могут занимать два слова («двадцати пяти»), поэтому длинный вариант
    пробуем первым: на «без двадцати пяти шесть» короткий прочитал бы «двадцати»
    и остался бы с «пяти» вместо часа.
    """
    if слова[i] != "без":
        return None
    for длина in (2, 1):
        конец = i + 1 + длина
        if конец >= len(слова):
            continue
        минуты = МИНУТЫ_РОДИТЕЛЬНЫЕ.get(" ".join(слова[i + 1:конец]))
        час = ЧАСЫ_КОЛИЧЕСТВЕННЫЕ.get(слова[конец])
        if минуты and час:
            return f"{час - 1:02d}:{60 - минуты:02d}", длина + 2
    return None


def _разговорное_время(слова):
    """Разговорное время в списке слов → «ЧЧ:ММ» одним токеном.

    Тот же приём, что у ИМЕНА_ЧАСОВ выше: переписываем в цифровую запись до
    основного разбора, и дальше «завтра в полдесятого» идёт по той же дороге,
    что «завтра в 9:30», — ни одна ветка ниже про разговорную форму не знает.
    Формы многословные («без пятнадцати десять»), поэтому проход по списку, а
    не подстановка по словарю.
    """
    итог, i = [], 0
    while i < len(слова):
        совпало = (_пол_часа(слова, i) or _четверть_часа(слова, i)
                   or _без_минут(слова, i))
        if совпало:
            токен, съедено = совпало
            итог.append(токен)
            i += съедено
        else:
            итог.append(слова[i])
            i += 1
    return итог


def _не_время(token):
    raise ValueError(f"не время: {token}")


def parse_time_part(token):
    """«14:00», «9.30», «18-45» → time. Не время — None, и это не ошибка:
    вызывающий просто поймёт, что времени в строке не было.

    Секунды принимаются и отбрасываются: точность продукта — минуты, но
    `str(datetime)` печатает время с секундами, а карточка отправляет то, что
    показал сервер, обратно как есть. Без этого «10:00:00» не опознаётся как
    время вообще, разбор проваливается в ISO-fallback на голую дату, и время
    у уже сохранённого шага молча пропадает при первом же сохранении карточки.
    """
    m = re.fullmatch(r"(\d{1,2})[:.\-](\d{2})(?::\d{2})?", token)
    if not m:
        return None
    часы, минуты = int(m.group(1)), int(m.group(2))
    if часы > 23 or минуты > 59:
        raise ValueError(f"не время: {token}")
    return dtime(часы, минуты)


def _час_или_время(token):
    """Хвост строки как время: и «9:30», и голый час «9»."""
    т = parse_time_part(token)
    if т is not None:
        return т
    if token.isdigit() and len(token) <= 2 and int(token) <= 23:
        return dtime(int(token), 0)
    return None


def _время_суток(t, слово):
    """«7 вечера» → 19:00, «12 ночи» → 00:00.

    Часы больше двенадцати не трогаем: «19 вечера» пишут редко, но кто написал,
    тот имел в виду ровно то, что написал, — доворачивать там нечего.
    """
    ч = t.hour
    if слово in ("дня", "вечера") and ч < 12:
        ч += 12
    elif слово in ("утра", "ночи") and ч == 12:
        ч = 0
    return dtime(ч, t.minute)


def parse_date_input(text, today, *, now=None):
    """Человеческий ввод → date или datetime. Здесь, а не в браузере.

    Форма и CLI обязаны понимать ввод одинаково, иначе появятся задачи, которые
    завелись через форму, но не читаются движком. Поэтому разбор один, а форма
    только показывает, во что он превратился.

    Понимает: 2026-08-18 · 18.08.2026 · 18.08 · сегодня · завтра · послезавтра ·
    +3 и «через 3 дня» · пн, вторник (ближайший такой день после сегодня) ·
    «15 марта» и «15 марта 2027» (месяц в любом падеже).

    Со временем: «завтра 14:00», «завтра в 9-30», «18.08 в 9», «+3 18:00»,
    «в 7 вечера», «завтра в полдень». Предлоги «в/во/на/до/после/к» выкидываются,
    время суток словом доворачивает час до суточного. Просто «14:00» — сегодня
    в это время. Без времени возвращается date, и шаг считается назначенным на
    весь день: рабочие часы для него считаются от начала дня.

    Разговорные формы времени: «полдесятого» и «половина десятого» (9:30),
    «четверть десятого» (9:15), «без пятнадцати десять» и «без четверти десять»
    (9:45), «без двадцати пяти шесть» (5:35). Разворачиваются в «ЧЧ:ММ» до
    основного разбора (`_разговорное_время`), поэтому складываются со всем
    остальным: «завтра в полдесятого вечера» — это 21:30 следующего дня.

    «Через час» / «через N часов» — от текущего момента, а не от полуночи
    `today`, поэтому требует именной аргумент `now`: без него `today` несёт
    только дату, часа в ней нет и отсчитывать не от чего. Вызовы, которым
    точный момент не нужен или недоступен (CLI с `--today`, предпросмотр
    шаблона), `now` не передают — фраза для них так и остаётся нераспознанной,
    с тем же `ValueError`, что и любой другой непонятный текст. Единственный
    вызывающий, у которого есть настоящее «сейчас», — окно контроля на ленте.
    """
    if text is None:
        return None
    s = str(text).strip().lower().replace("ё", "е")
    if not s:
        return None

    if now is not None:
        часовой = re.fullmatch(r"через\s+(\d+)?\s*час\w*", s)
        if часовой:
            return now + timedelta(hours=int(часовой.group(1) or 1))

    слова = [ИМЕНА_ЧАСОВ.get(w, w) for w in s.split() if w not in СВЯЗКИ]
    if not слова:
        raise ValueError(f"не дата: {text!r}")
    слова = _разговорное_время(слова)
    s = " ".join(слова)

    # «через 3 дня» — то же самое, что «+3», просто длиннее написано. Хвост
    # («18:00» после «дня») не теряем — он уйдёт в обычный разбор времени ниже.
    через = re.match(r"^через\s+(\d+)\s*д(?:ень|ня|ней)?\b\s*(.*)$", s)
    if через:
        s = f"+{через.group(1)} {через.group(2)}".strip()

    # «вечера» относится к часу перед собой, поэтому слово снимаем заранее:
    # дальше время разбирается обычным путём, как если бы его написали цифрами.
    # Порядок важен — «дня» из «через 3 дня» разобрано выше и сюда не доходит.
    части = s.split()
    суточное = None
    if len(части) > 1 and части[-1] in ЧАСТИ_СУТОК:
        суточное = части.pop()
        s = " ".join(части)

    # Время отделяем до всего остального: «18.08 09:30» — это дата и время, а не
    # два непонятных числа.
    if len(части) > 1:
        часть_времени = _час_или_время(части[-1])
        if часть_времени is not None:
            if суточное:
                часть_времени = _время_суток(часть_времени, суточное)
            день = parse_date_input(" ".join(части[:-1]), today)
            if день is None:
                raise ValueError(f"есть время, но нет даты: {text!r}")
            return datetime.combine(as_date(день), часть_времени)
    else:
        # Одно слово. Часом его считаем, только когда это однозначно час: стоит
        # двоеточие, приписано время суток или это короткое число вроде «9».
        # Голое «18» не путаем с датой: «18.08» ловится дальше своим форматом.
        if суточное or ":" in s:
            часы = _час_или_время(s) or _не_время(s)
            return datetime.combine(
                today, _время_суток(часы, суточное) if суточное else часы)
        if s.isdigit() and len(s) <= 2 and int(s) <= 23:
            return datetime.combine(today, dtime(int(s), 0))

    if s in ОТНОСИТЕЛЬНЫЕ:
        return today + timedelta(days=ОТНОСИТЕЛЬНЫЕ[s])

    if s.startswith("+") and s[1:].isdigit():
        return today + timedelta(days=int(s[1:]))

    день = ДНИ_НЕДЕЛИ.get(s)
    if день is not None:
        вперёд = (день - today.weekday()) % 7 or 7  # «пн» в понедельник — следующий
        return today + timedelta(days=вперёд)

    # «15 марта», «15 марта 2027». Голый месяц без числа («в мае») намеренно не
    # проходит: у шага одна дата контроля, а «май» — это тридцать один вариант,
    # и выбирать за человека первое число значит тихо соврать.
    m = re.fullmatch(r"(\d{1,2})\s+([а-я]+)(?:\s+(\d{4}))?", s)
    if m and _месяц(m.group(2)):
        д, месяц, год = int(m.group(1)), _месяц(m.group(2)), m.group(3)
        if год:
            return date(int(год), месяц, д)
        дата = date(today.year, месяц, д)
        # «15 марта» в августе — это следующий март, та же логика, что у «18.08»
        return дата if дата >= today else date(today.year + 1, месяц, д)

    m = re.fullmatch(r"(\d{1,2})[.\-/](\d{1,2})(?:[.\-/](\d{2}|\d{4}))?", s)
    if m:
        д, мес, год = int(m.group(1)), int(m.group(2)), m.group(3)
        if год is None:
            # «9-30» — не тридцатый месяц, а половина десятого. Если датой строка
            # не читается, пробуем прочитать её временем, а не роняем разбор.
            if мес > 12:
                return datetime.combine(today, parse_time_part(s) or _не_время(s))
            дата = date(today.year, мес, д)
            # «18.08» в сентябре — это следующий год, а не прошедшая дата
            return дата if дата >= today else date(today.year + 1, мес, д)
        год = int(год)
        return date(год + 2000 if год < 100 else год, мес, д)

    return as_date(s)  # ISO и всё, что понимает datetime.fromisoformat


# После переезда на SQLite название задачи больше не имя файла — ограничение
# сохранено, но теперь по другой причине. `|` занят синтаксисом piped-ссылок
# базы знаний (`[[Название|как написано]]`), `\/:*?"<>` в названии задачи
# нечитаемы и мешали бы будущим выгрузкам (Excel режет такие символы в именах
# листов). Осознанное продуктовое правило, не отпечаток файловой системы.
FORBIDDEN_IN_NAME = set('\\/:*?"<>|')

# Экран не резиновый: 120 символов — предел, после которого название в ленте
# и в списке шагов перестаёт помещаться в строку, а не ограничение файловой
# системы (тем оно и было раньше, для markdown-файлов).
MAX_TITLE = 120


def _parse_optional_date(raw, today, поле, errors):
    """Разобрать необязательную дату, добавить ошибку в список при провале.

    Общий кусок между валидацией контрольной даты и даты начала — раньше жил
    только внутри `validate_new_task`, теперь нужен в двух местах и разойтись
    им нельзя: разное сообщение об ошибке на одну и ту же дату сбивает с толку.
    """
    raw = (raw or "").strip()
    if not raw:
        return None
    try:
        return parse_date_input(raw, today)
    except (ValueError, TypeError):
        errors.append({"field": поле,
                       "error": "Дату не понял. Можно: 18.08 · 15 марта · завтра · +3 · пн · полдесятого"})
        return None


def default_step_start(предыдущий_контроль, старт_задачи):
    """Дата начала шага по умолчанию — раздел 6.3.2 ТЗ: контроль предыдущего
    шага, а для первого шага дата начала задачи."""
    return предыдущий_контроль or старт_задачи


MODES = ("par", "seq")


def resolve_steps(steps_data, старт_задачи, today, старые=None, prefix="steps",
                  предыдущий=None, mode="seq"):
    """Разобрать шаги и подставить дефолт даты начала — один проход, которым
    пользуются и проверка, и запись, и создание, и правка.

    Раньше дефолт вычислялся заново в `build_task`, отдельно от `validate_new_task`:
    проверка смотрела только на то, что пришло в запросе, и дефолт мог обогнать
    control_date уже после проверки. Здесь дефолт и проверка смотрят на одни
    и те же значения.

    Шаг с непустым списком `steps` — группа: у неё режим ("par" по умолчанию,
    "seq" для подцепочки), дат нет, дети разбираются рекурсивно. Дефолт даты
    начала листа — по последовательной цепочке: контроль предыдущего элемента,
    у группы это максимум контролей её поддерева. Внутри параллельной группы
    цепочки нет: каждый ребёнок стартует от точки входа группы.

    `старые` — режим правки: для листа с известным `id` без явной даты начала
    берётся сохранённое, а не дефолт по новому порядку. Без этого чистая
    перестановка шагов в карточке упиралась бы в «контроль раньше начала».

    Возвращает дерево словарей: title, start, control, note, id, mode,
    children, errors, явный_старт, поле, предыдущий (для мягких предупреждений).
    """
    resolved = []
    for i, step in enumerate(steps_data or []):
        поле = f"{prefix}.{i}"
        errors = []
        if not (step.get("title") or "").strip():
            errors.append({"field": f"{поле}.title", "error": "Название шага обязательно"})
        дети_данные = step.get("steps") or []
        node = {"title": step.get("title"), "note": step.get("note"),
                "id": step.get("id"), "errors": errors, "children": [],
                "mode": None, "start": None, "control": None,
                "явный_старт": False, "поле": поле, "предыдущий": предыдущий}
        if дети_данные or step.get("mode"):
            режим = step.get("mode") or "par"
            if режим not in MODES:
                errors.append({"field": f"{поле}.mode",
                               "error": "Режим группы — par или seq"})
                режим = "par"
            node["mode"] = режим
            if not дети_данные:
                errors.append({"field": f"{поле}.steps",
                               "error": "В группе нужен хотя бы один подшаг"})
            for k in ("control_date", "start_date"):
                if step.get(k):
                    errors.append({"field": f"{поле}.{k}",
                                   "error": "Даты ставятся подшагам, не группе"})
            node["children"] = resolve_steps(
                дети_данные, старт_задачи, today, старые,
                prefix=f"{поле}.steps", предыдущий=предыдущий, mode=режим)
            финиш = _финиш_узла(node)
        else:
            control = _parse_optional_date(step.get("control_date"), today,
                                           f"{поле}.control_date", errors)
            явный_старт = _parse_optional_date(step.get("start_date"), today,
                                               f"{поле}.start_date", errors)
            сохранённый = None
            if старые is not None and step.get("id") in старые:
                сохранённый = as_date(старые[step["id"]].get("start_date"))
            start = явный_старт or сохранённый or default_step_start(предыдущий,
                                                                     старт_задачи)
            # Жёсткая проверка из раздела 6.3.3 ТЗ: контроль раньше, чем шаг можно
            # начать, бессмысленен как дата — блокирует сохранение. Сравниваем с
            # итоговым start (явным или дефолтным), а не только с введённым.
            if start and control and as_date(control) < as_date(start):
                errors.append({"field": f"{поле}.control_date",
                               "error": "Контроль раньше даты начала шага"})
            node.update(start=start, control=control,
                        явный_старт=bool(явный_старт))
            финиш = control
        resolved.append(node)
        if mode != "par" and финиш:
            предыдущий = финиш
    return resolved


def _финиш_узла(node):
    """Когда элемент цепочки «кончается» для дефолта следующего: у листа это
    его контроль, у группы — самый поздний контроль поддерева. None, если дат
    в поддереве нет вовсе."""
    даты = [node["control"]] if node["control"] else []
    даты += [f for f in (_финиш_узла(c) for c in node["children"]) if f]
    return max(даты, key=as_date) if даты else None


def walk_resolved(nodes):
    """Дерево resolve_steps плоским потоком, глубина-первым порядком."""
    for n in nodes:
        yield n
        yield from walk_resolved(n["children"])


def validate_new_task(data, existing_names, today):
    """Проверка задачи до записи. Возвращает список ошибок по полям — тех,
    что блокируют сохранение. Мягкие предупреждения — отдельно, в `soft_warnings`.

    Отдельно от формы намеренно: правила должны быть в одном месте, иначе форма
    и CLI разойдутся, и в вольт попадёт то, что движок потом не прочитает.
    Ошибки возвращаются списком, а не первым попавшимся исключением, — форме надо
    подсветить все проблемные поля разом, а не гонять человека по кругу.
    """
    errors = []

    title = (data.get("title") or "").strip()
    if not title:
        errors.append({"field": "title", "error": "Название задачи обязательно"})
    elif len(title) > MAX_TITLE:
        errors.append({"field": "title",
                       "error": f"Название длиннее {MAX_TITLE} символов"})
    elif set(title) & FORBIDDEN_IN_NAME:
        плохие = "".join(sorted(set(title) & FORBIDDEN_IN_NAME))
        errors.append({"field": "title",
                       "error": f"В названии нельзя символы {плохие}"})
    elif title.lower() in {n.lower() for n in existing_names}:
        errors.append({"field": "title",
                       "error": "Задача с таким названием уже есть"})
    elif title != title.strip(". "):
        errors.append({"field": "title",
                       "error": "Название не должно кончаться точкой или пробелом"})

    старт_задачи = _parse_optional_date(data.get("start_date"), today,
                                        "start_date", errors) or today

    steps = data.get("steps") or []
    if not steps:
        errors.append({"field": "steps", "error": "Нужен хотя бы один шаг"})
    for r in walk_resolved(resolve_steps(steps, старт_задачи, today)):
        errors += r["errors"]
    return errors


def soft_warnings(data, today):
    """Мягкие предупреждения из раздела 6.3.3 ТЗ: сохранить можно, но человек
    должен увидеть, что даты выглядят подозрительно.

    Не блокируют запись, поэтому отдельная функция, а не часть `validate_new_task`:
    смешивать в одном списке то, что останавливает сохранение, с тем, что просто
    предупреждает, заставило бы форму гадать, какая ошибка какая.

    Сравнение идёт по явно введённой дате начала, не по дефолтной: дефолт равен
    как раз тому, с чем его сравнивают (концу задачи или предыдущему шагу), и
    строгое «меньше» на них никогда не сработает — предупреждать не о чем.
    """
    warnings = []
    старт_задачи = _parse_optional_date(data.get("start_date"), today, None, []) or today
    for r in walk_resolved(resolve_steps(data.get("steps") or [], старт_задачи, today)):
        if not r["явный_старт"]:
            continue
        if as_date(r["start"]) < as_date(старт_задачи):
            warnings.append({"field": f'{r["поле"]}.start_date',
                             "warning": "Шаг начинается раньше даты начала задачи"})
        if r["предыдущий"] and as_date(r["start"]) < as_date(r["предыдущий"]):
            warnings.append({"field": f'{r["поле"]}.start_date',
                             "warning": "Шаг начинается раньше, чем закончится "
                                       "предыдущий"})
    return warnings


def build_task(data, today):
    """Данные формы → frontmatter задачи. Без записи на диск.

    Идентификаторы шагов раздаёт движок, а не форма: они должны быть плотными и
    по порядку, иначе `done <задача> 3` будет попадать не туда.
    """
    старт_задачи = _parse_optional_date(data.get("start_date"), today, None, []) or today
    steps = []

    def добавить(nodes, parent):
        for r in nodes:
            sid = len(steps) + 1
            steps.append({
                "id": sid,
                "title": r["title"].strip(),
                # У группы статус смысла не несёт (закрытие вычисляется из
                # детей), но форма записи шага одна на всех — колонка NOT NULL.
                "status": OPEN,
                "start_date": r["start"],
                "control_date": r["control"],
                "completed_date": None,
                "note": (r["note"] or "").strip() or None,
                "parent": parent,
                "mode": r["mode"],
                "log": [],
            })
            добавить(r["children"], sid)

    добавить(resolve_steps(data.get("steps") or [], старт_задачи, today), None)
    tags = [t.strip() for t in (data.get("tags") or []) if t and t.strip()]
    meta = {
        "schema": SCHEMA,
        "type": "task",
        "title": data["title"].strip(),
        "created": today,
        "start_date": старт_задачи,
        "tags": tags,
        "steps": steps,
    }
    return meta


def cmd_create(args, today):
    """Завести задачу. Единственный путь создания — и из формы, и из CLI.

    JSON на входе: {"title": "...", "tags": [...], "steps": [{"title": "...",
    "control_date": "2026-08-20"}], "body": "..."}
    """
    raw = sys.stdin.read() if args.json == "-" else args.json
    try:
        data = json.loads(raw)
    except json.JSONDecodeError as e:
        return {"ok": False, "errors": [{"field": None, "error": f"битый JSON: {e}"}]}

    existing = [t["path"].stem for t in load_tasks()]
    errors = validate_new_task(data, existing, today)
    if errors:
        return {"ok": False, "errors": errors}

    meta = build_task(data, today)
    # `path` — раньше был предвычисленный путь к файлу, теперь задачи ещё нет
    # в БД, значит нет и id. save() увидит path=None, заведёт новую строку и
    # сам подставит сюда TaskRef с настоящим id.
    task = {"path": None, "meta": meta, "body": (data.get("body") or "").strip() + "\n"}
    try:
        save(task, today)
    except store.DuplicateTitle:
        return {"ok": False, "errors": [{"field": "title", "error": "Задача с таким названием уже есть"}]}
    return {"ok": True, "task": task["path"].stem,
            "steps": len(meta["steps"]), "status": task["meta"]["status"]}


# --- правка существующей задачи --------------------------------------------
#
# Отдельно от создания. При правке шаги приходят с уже известными `id`, и эти
# id обязаны пережить редактирование: на них ссылаются `done`/`notdone`/
# `defer`/`fail`/`skip`, и переезд с 1..N при каждом сохранении раскидал бы
# отметки не по тем шагам.
#
# Статус, дата закрытия и журнал шага правкой не трогаются никогда — это поле
# зоны четырёх команд перехода, а не карточки. Карточка меняет только то, что
# заказчик видит как метаданные: заголовок, даты, заметку, порядок, состав.

def validate_task_edit(task, data, existing_names, today):
    """Проверка правки — со своими правилами дат (см. `resolve_steps_for_edit`),
    а не `validate_new_task`: та не знает про сохранённые даты существующих
    шагов и на чистой перестановке без единой правки дат ошибалась бы сама.
    """
    errors = []
    title = (data.get("title") or "").strip()
    свои = {n for n in existing_names if n.lower() != task["path"].stem.lower()}
    if not title:
        errors.append({"field": "title", "error": "Название задачи обязательно"})
    elif len(title) > MAX_TITLE:
        errors.append({"field": "title",
                       "error": f"Название длиннее {MAX_TITLE} символов"})
    elif set(title) & FORBIDDEN_IN_NAME:
        плохие = "".join(sorted(set(title) & FORBIDDEN_IN_NAME))
        errors.append({"field": "title",
                       "error": f"В названии нельзя символы {плохие}"})
    elif title.lower() in {n.lower() for n in свои}:
        errors.append({"field": "title",
                       "error": "Задача с таким названием уже есть"})
    elif title != title.strip(". "):
        errors.append({"field": "title",
                       "error": "Название не должно кончаться точкой или пробелом"})

    старт_задачи = _parse_optional_date(data.get("start_date"), today, "start_date",
                                        errors) or as_date(task["meta"].get("start_date")) \
        or today
    старые = {s["id"]: s for s in steps_of(task)}

    steps = data.get("steps") or []
    if not steps:
        errors.append({"field": "steps", "error": "Нужен хотя бы один шаг"})
    for r in walk_resolved(resolve_steps(steps, старт_задачи, today, старые=старые)):
        errors += r["errors"]
    return errors


def apply_task_edit(task, data, today):
    """Переписать метаданные задачи по данным карточки. Возвращает список
    id шагов, которые пропали из данных без явного «снять» — молчаливая потеря
    шага хуже, чем отказ сохранить.
    """
    meta = task["meta"]
    старые = {s["id"]: s for s in steps_of(task)}
    старт_задачи = _parse_optional_date(data.get("start_date"), today, None, []) or \
        as_date(meta.get("start_date")) or today

    следующий_id = max([s["id"] for s in старые.values()], default=0) + 1
    новые, увиденные = [], set()

    def добавить(nodes, parent):
        nonlocal следующий_id
        for r in nodes:
            если_старый = r["id"] is not None and r["id"] in старые
            if если_старый:
                шаг = dict(старые[r["id"]])  # статус/completed_date/log копируются как есть
                увиденные.add(r["id"])
            else:
                # Тот же порядок полей, что у build_task, — иначе новый шаг в файле
                # выглядит написанным другой рукой, хотя человеку разницы нет.
                шаг = {"id": следующий_id, "title": None, "status": OPEN,
                       "start_date": None, "control_date": None,
                       "completed_date": None, "note": None, "parent": None,
                       "mode": None, "log": []}
                следующий_id += 1
            шаг["title"] = r["title"].strip()
            шаг["start_date"] = r["start"]
            шаг["control_date"] = r["control"]
            шаг["note"] = (r["note"] or "").strip() or None
            # Родитель и режим переписываются и у старых шагов: карточка могла
            # перетащить шаг в группу или обратно, это правка структуры, а не
            # статуса. Статус и журнал при этом не трогаются.
            шаг["parent"] = parent
            шаг["mode"] = r["mode"]
            новые.append(шаг)
            добавить(r["children"], шаг["id"])

    добавить(resolve_steps(data.get("steps") or [], старт_задачи, today, старые=старые),
             None)

    пропали = [sid for sid in старые if sid not in увиденные]

    meta["title"] = data.get("title", meta["title"]).strip()
    meta["start_date"] = старт_задачи
    if "tags" in data:
        meta["tags"] = [t.strip() for t in (data.get("tags") or []) if t and t.strip()]
    meta["steps"] = новые
    if "body" in data:
        task["body"] = (data.get("body") or "").strip() + "\n"
    return пропали


def cmd_update(args, today):
    """Правка задачи из карточки: заголовок, даты, теги, заметка, состав и
    порядок шагов. Статусы шагов через `done`/`notdone`/`defer`/`fail`/`skip`.
    """
    raw = sys.stdin.read() if args.json == "-" else args.json
    try:
        data = json.loads(raw)
    except json.JSONDecodeError as e:
        return {"ok": False, "errors": [{"field": None, "error": f"битый JSON: {e}"}]}

    task = find_task(args.task)
    прежнее_имя = task["path"].stem
    existing = [t["path"].stem for t in load_tasks()]
    errors = validate_task_edit(task, data, existing, today)
    if errors:
        return {"ok": False, "errors": errors}

    пропали = apply_task_edit(task, data, today)
    if пропали and not getattr(args, "force", False):
        # Шаг исчез из данных без явного «снять». Скорее всего баг интерфейса
        # или случайное перетаскивание мимо списка — молча терять историю шага
        # нельзя, поэтому здесь отказ, а не тихое удаление.
        return {"ok": False, "errors": [{
            "field": "steps",
            "error": f"Шаг {sid} пропал из данных — сначала «снять», не убирать так"}
            for sid in пропали]}

    # Переименование раньше значило перенос файла — os.replace после записи,
    # отдельная проверка «файл уже существует» до неё. У задачи-строки id не
    # меняется от смены title, так что переименование — то же самое save(),
    # что и любая другая правка; `validate_task_edit` уже проверила название
    # на совпадение с другими задачами, UNIQUE(title) в БД — подстраховка от
    # гонки, а не источник этой проверки.
    try:
        save(task, today)
    except store.DuplicateTitle:
        return {"ok": False, "errors": [{"field": "title", "error": "Задача с таким названием уже есть"}]}

    # Строку индекса адресует название, поэтому переименованная задача оставила
    # бы позади себя старую: она находилась бы по прежнему слову и вела в
    # никуда. `save` уже записал новую — снимаем только прежнюю.
    if task["path"].stem != прежнее_имя:
        get_store().search_forget("task", прежнее_имя)

    return {"ok": True, "task": task["path"].stem, "status": task["meta"]["status"],
            "steps": len(task["meta"]["steps"])}


def cmd_cancel(args, today):
    """Отменить задачу целиком — раздел 6.3 ТЗ, кнопка «Отменить задачу».

    Не «снять» (это про один шаг) и не удаление (файл и история остаются).
    Отменённая задача перестаёт быть просроченной или ждущей: её статус
    вычисляется первым делом в `task_status`, раньше любого правила про шаги.
    """
    task = find_task(args.task)
    if task["meta"].get("cancelled"):
        sys.exit(f"задача «{task['path'].stem}» уже отменена")
    task["meta"]["cancelled"] = True
    task["meta"]["cancelled_reason"] = (getattr(args, "reason", None) or "").strip() or None
    save(task, today)
    return {"ok": True, "task": task["path"].stem, "status": task["meta"]["status"]}


def cmd_delete(args, today):
    """Удалить задачу насовсем. Подтверждение — дело интерфейса, не движка:
    здесь только сам необратимый шаг."""
    task = find_task(args.task)
    склад = get_store()
    склад.delete_task(task["path"].id)
    # Из индекса тоже: иначе удалённая задача продолжает находиться поиском, и
    # клик по ней ведёт в никуда.
    склад.search_forget("task", task["path"].stem)
    return {"ok": True, "task": task["path"].stem, "deleted": True}


def cmd_reopen(args, today):
    """Отменить закрытие последнего сделанного шага — раздел 6.3.5 ТЗ:
    задача закрывается автоматически, «показывается подтверждение с
    возможностью отменить». Открывает конкретный шаг, не «последний вообще»:
    порядок закрытия и порядок в списке могут не совпасть при провале/снятии
    более раннего шага.
    """
    task = find_task(args.task)
    step = get_step(task, args.step)
    if step.get("status") != DONE:
        sys.exit(f"шаг {args.step} не был сделан (сейчас: {step.get('status')})")
    step["status"] = OPEN
    step["completed_date"] = None
    log_event(step, "reopened", today)
    save(task, today)
    return {"ok": True, "task": task["path"].stem, "step": args.step,
            "task_status": task["meta"]["status"]}


def _recurrence_view(шаблон):
    """Правило повторения с человеческой подписью — для карточки шаблона.
    Подпись считает `recurrence.describe`, а не морда: тот же принцип, что и
    везде — оболочка не пересказывает правило словами сама."""
    правило = шаблон.get("recurrence")
    if not правило:
        return None
    return {**правило, "description": rec.describe(
        {k: v for k, v in правило.items() if k != "anchor"})}


def cmd_templates(args, today):
    """Список шаблонов. Отдаём с предпросмотром на сегодня: заказчику надо видеть,
    какие даты получатся, а не только имена."""
    склад = tpl.JsonStore(VAULT)
    старт = parse_date_input(args.start, today) if getattr(args, "start", None) else today
    из_даты = as_date(старт)
    файлы = get_store()
    out = []
    for шаблон in склад.all():
        out.append({
            "name": шаблон["name"],
            "tags": шаблон.get("tags") or [],
            "steps": len(шаблон.get("steps") or []),
            "attachments": len(файлы.list_attachments("template", шаблон["name"])),
            "preview": tpl.preview(шаблон, из_даты),
            "recurrence": _recurrence_view(шаблон),
        })
    return {"templates": out, "count": len(out), "start": из_даты.isoformat()}


def cmd_template_preview(args, today):
    """Какие даты дадут шаги ещё не сохранённого шаблона. Раздел 5.6 ТЗ.

    Форма шаблона показывает сдвиги в днях, а днями человек не думает: «+10»
    выглядит правдоподобно ровно до того момента, когда попадает на праздники.
    Ту же роль играет `/api/parse-date` в форме задачи — считает и проверяет
    ядро, страница показывает ответ.

    Название здесь не проверяется намеренно: даты от него не зависят, а человек
    набирает шаги раньше, чем придумывает имя, и ругаться на пустое поле в
    предпросмотре незачем — на сохранении оно и так не пройдёт.
    """
    raw = sys.stdin.read() if args.json == "-" else args.json
    try:
        data = json.loads(raw)
    except json.JSONDecodeError as e:
        return {"ok": False, "errors": [{"field": None, "error": f"битый JSON: {e}"}]}

    старт = as_date(parse_date_input(args.start, today)) \
        if getattr(args, "start", None) else today
    пробный = {**data, "name": (data.get("name") or "").strip() or "—", "recurrence": None}
    ошибки = [e for e in tpl.validate_template(пробный)
              if not str(e.get("field") or "").startswith("name")]
    if ошибки:
        return {"ok": False, "errors": ошибки}
    return {"ok": True, "start": старт.isoformat(),
            "start_text": tpl.human_moment(старт),
            "steps": tpl.preview(пробный, старт)}


def cmd_recurrence_preview(args, today):
    """Проверить и описать правило без сохранения — живая подпись в форме,
    та же роль, что у `/api/parse-date` для одиночной даты."""
    if not getattr(args, "anchor", None):
        return {"ok": False, "errors": [{"field": "recurrence.anchor",
                                         "error": "Нужна дата, от которой считать первый цикл"}]}
    try:
        якорь = as_date(parse_date_input(args.anchor, today))
    except (ValueError, TypeError):
        return {"ok": False, "errors": [{"field": "recurrence.anchor",
                                         "error": "Дату не понял, нужен формат 2026-08-18"}]}
    try:
        правило = json.loads(args.rule) if isinstance(args.rule, str) else (args.rule or {})
    except json.JSONDecodeError as e:
        return {"ok": False, "errors": [{"field": "recurrence", "error": f"битый JSON: {e}"}]}

    ошибки = rec.validate_rule(правило, start=якорь)
    if ошибки:
        return {"ok": False, "errors": [
            {"field": f"recurrence.{e['field']}" if e.get("field") else "recurrence",
             "error": e["error"]} for e in ошибки]}

    return {"ok": True, "description": rec.describe(правило), "anchor": якорь.isoformat(),
            "preview": [{"date": s["date"].isoformat(), "text": s["text"]}
                       for s in rec.preview(правило, якорь, count=5)]}


def cmd_recurrence_parse(args, today):
    """Текст «каждый вторник» → правило повторения. Требование R23.

    Разбирает `recurrence.parse_text` — детерминированно, без модели: это
    словарь на два десятка слов, а не связная речь, и по границе из CLAUDE.md
    он на нашей стороне. Заодно правило не зависит от сети, а форма может
    дёргать разбор на каждое нажатие клавиши.

    Наружу уходит и правило, и его подпись: разобрав текст, надо показать
    человеку, как система его поняла, — «каждую неделю по вторникам» рядом с
    ближайшими датами. Иначе разбор молча съедает опечатку.
    """
    try:
        правило = rec.parse_text(getattr(args, "text", None))
    except rec.RuleError as e:
        return {"ok": False, "errors": e.errors}
    if правило.get("until") is not None:
        правило["until"] = правило["until"].isoformat()
    return {"ok": True, "rule": правило, "description": rec.describe(правило)}


def cmd_set_recurrence(args, today):
    """Прикрепить или снять правило повторения с шаблона.

    Идёт через `Store.save` целиком, а не отдельным полем: у шаблона один путь
    записи, тот же, что у формы шагов, — иначе однажды разойдутся форматом.
    """
    склад = tpl.JsonStore(VAULT)
    шаблон = склад.get(args.name)
    if not шаблон:
        return {"ok": False, "errors": [{"field": "name",
                                         "error": f"нет шаблона «{args.name}»"}]}
    данные = dict(шаблон)
    if getattr(args, "clear", False):
        данные["recurrence"] = None
    else:
        данные["recurrence"] = args.rule if isinstance(args.rule, dict) else json.loads(args.rule)
    try:
        обновлённый = склад.save(данные)
    except tpl.TemplateError as e:
        return {"ok": False, "errors": getattr(e, "errors", [{"field": None, "error": str(e)}])}
    return {"ok": True, "template": обновлённый["name"],
            "recurrence": _recurrence_view(обновлённый)}


def cmd_save_template(args, today):
    """Создать шаблон с нуля или переписать существующий целиком.

    Раньше единственным путём завести шаблон было «сохранить как» у уже готовой
    задачи (`cmd_template_from_task`) — с нуля собрать было нечем, хотя
    `tpl.Store.save` шаблон без задачи-источника принимает и так. Здесь тот же
    вызов, что там, только данные приходят прямо от формы или из CLI, а не из
    развёрнутых дат существующей задачи.

    JSON на входе: {"name": "...", "tags": [...], "steps": [{"title": "...",
    "offset_days": 0, "time_of_day": "14:00"}], "body": "..."}
    """
    raw = sys.stdin.read() if args.json == "-" else args.json
    try:
        data = json.loads(raw)
    except json.JSONDecodeError as e:
        return {"ok": False, "errors": [{"field": None, "error": f"битый JSON: {e}"}]}

    склад = tpl.JsonStore(VAULT)
    try:
        шаблон = склад.save(data)
    except tpl.TemplateError as e:
        return {"ok": False, "errors": getattr(e, "errors", [{"field": None, "error": str(e)}])}
    return {"ok": True, "template": шаблон["name"], "steps": len(шаблон.get("steps") or [])}


def _create_task_from_data(данные, today, existing=None):
    """Общий путь записи новой задачи — из формы, из шаблона, из повторения.

    Один путь, а не три копии: второй писатель рано или поздно разойдётся с
    первым в мелочи вроде порядка полей или сводки. `existing` передают, когда
    список задач уже прочитан вызывающим (движок повторений создаёт несколько
    задач подряд, и читать вольт заново перед каждой — лишний проход по файлам).
    """
    существующие = existing if existing is not None else [t["path"].stem for t in load_tasks()]
    errors = validate_new_task(данные, существующие, today)
    if errors:
        return None, errors

    meta = build_task(данные, today)
    # Происхождение проставляется до записи и только здесь: `build_task` про
    # шаблоны не знает и знать не должен, а `save` пишет то, что в meta.
    if данные.get("template_name"):
        meta["template_name"] = данные["template_name"]
        meta["cycle_key"] = данные.get("cycle_key")
    задача = {"path": None, "meta": meta, "body": (данные.get("body") or "").strip() + "\n"}
    try:
        save(задача, today)
    except store.DuplicateTitle:
        return None, [{"field": "title", "error": "Задача с таким названием уже есть"}]
    return задача, None


def copy_template_attachments(template_name, task_name, today):
    """Перенести файлы шаблона на заведённую из него задачу. Возвращает,
    сколько перенесено.

    Копируются строки в `attachments`, а не байты: файл на диске адресуется
    своим sha256 (`attachments.save`), поэтому вторая ссылка на ту же картинку
    ничего не пишет на диск и ничего не весит. Без этого шага фото на шаблоне
    остаётся украшением карточки: человек делает задачу, а схема, ради которой
    файл прикрепляли, лежит там, куда он в этот момент не смотрит.

    Подпись и имя файла переносятся как есть, дата ставится сегодняшняя — это
    дата появления файла у задачи, а не у шаблона.
    """
    склад = get_store()
    перенесено = 0
    for r in склад.list_attachments("template", template_name):
        склад.add_attachment("task", task_name, r["sha256"], r["filename"],
                             r["mime"], r["bytes"], r["caption"], today)
        перенесено += 1
    return перенесено


def cmd_from_template(args, today):
    """Завести задачу из шаблона."""
    склад = tpl.JsonStore(VAULT)
    шаблон = склад.get(args.name)
    if not шаблон:
        return {"ok": False, "errors": [{"field": "name",
                                         "error": f"нет шаблона «{args.name}»"}]}
    старт = as_date(parse_date_input(args.start, today)) if args.start else today
    данные = tpl.expand(шаблон, старт, title=args.title)

    задача, errors = _create_task_from_data(данные, today)
    if errors:
        return {"ok": False, "errors": errors}
    файлы = copy_template_attachments(шаблон["name"], задача["path"].stem, today)
    return {"ok": True, "task": задача["path"].stem, "template": шаблон["name"],
            "steps": len(задача["meta"]["steps"]), "attachments": файлы,
            "status": задача["meta"]["status"]}


# --- повторения --------------------------------------------------------

def recurrence_state_path(vault=None):
    """Журнал повторений лежит рядом с вольтом, но не в нём: это отметки «какой
    цикл был последним», а не данные заказчика. Тот же приём, что у файла
    доставки в notify.py — потеря файла означает лишний повтор, а не потерю
    задачи."""
    return Path(vault or VAULT) / ".повторения.json"


def load_recurrence_state():
    path = recurrence_state_path()
    if not path.is_file():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (ValueError, OSError):
        # Битый журнал — не повод падать: хуже пропустить проверку блокировки
        # один раз, чем перестать заводить задачи по всем правилам разом.
        return {}


def save_recurrence_state(state):
    path = recurrence_state_path()
    fd, tmp = tempfile.mkstemp(dir=str(path.parent), suffix=".tmp")
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            json.dump(state, f, ensure_ascii=False, indent=1, default=str)
        os.replace(tmp, path)
    except BaseException:
        Path(tmp).unlink(missing_ok=True)
        raise


def recurring_title(name, cycle_date):
    """Имя автосозданной задачи. Голое имя шаблона совпало бы с прошлым циклом —
    ровно та коллизия, что при ручном разворачивании ловит понятную ошибку
    в форме, а здесь заведение идёт без человека и споткнуться не о что."""
    return f"{name} — {cycle_date:%d.%m.%Y}"


def _cycle_closed(task_name, today):
    """Закрыт ли цикл — по статусу задачи, которую он породил.

    Файл мог исчезнуть: заказчик вправе удалить задачу руками. Отсутствие
    считаем закрытием, а не блокировкой навсегда — иначе удалённая вручную
    задача остановила бы правило насовсем, и это тише любой ошибки.
    """
    for задача in load_tasks():
        if задача["path"].stem == task_name:
            return task_status(задача, today) == "done"
    return True


def cmd_recur(args, today):
    """Прогнать шаблоны с правилом повторения: создать очередной цикл или
    записать пропуск. Раздел 5.12 ТЗ.

    Идемпотентно в границах контракта: незакрытый цикл при повторном вызове в
    тот же день снова даёт пропуск, а не вторую задачу — `due_cycles` сам не
    продвигает журнал, пока предыдущий цикл не закрыт.
    """
    склад = tpl.JsonStore(VAULT)
    state = load_recurrence_state()
    work = worktime.settings()
    задачи_кэш = [t["path"].stem for t in load_tasks()]

    отчёт = []
    for шаблон in склад.all():
        правило = шаблон.get("recurrence")
        if not правило:
            continue
        имя = шаблон["name"]
        запись = state.get(имя) or {}
        предыдущий = None
        if запись.get("previous"):
            предыдущий = dict(запись["previous"])
            задача_цикла = предыдущий.pop("task", None)
            if задача_цикла:
                предыдущий["closed"] = _cycle_closed(задача_цикла, today)

        якорь = as_date(правило["anchor"])
        сила = bool(getattr(args, "force", False)) and getattr(args, "name", None) == имя
        try:
            решения = rec.due_cycles(
                {k: v for k, v in правило.items() if k != "anchor"}, якорь, today,
                previous=предыдущий, work=work, force=сила,
                limit=getattr(args, "limit", None) or 12)
        except rec.RuleError as e:
            отчёт.append({"template": имя, "errors": e.errors, "created": [], "skipped": []})
            continue

        # Статус закрытия старого цикла пересчитывается заново на каждом вызове
        # из фактического состояния задачи (см. выше), а не хранится, — поэтому
        # если новых циклов в этом прогоне не появилось, запись про «previous»
        # трогать не нужно вовсе: она и так будет пересчитана в следующий раз.
        созданы, пропущены, сбой = [], [], None
        for решение in решения:
            if решение["action"] == "skip":
                пропущены.append({"date": решение["date"].isoformat(),
                                  "message": решение["message"]})
                continue
            title = recurring_title(имя, решение["date"])
            данные = tpl.expand(шаблон, решение["date"], title=title)
            # Откуда задача взялась — в колонки, а не в разбор названия потом.
            # `решение["key"]` это `recurrence.cycle_key`, тот же ключ, которым
            # журнал повторений отличает уже записанный цикл от нового.
            данные["template_name"] = имя
            данные["cycle_key"] = решение["key"]
            задача, errors = _create_task_from_data(данные, today, existing=задачи_кэш)
            if errors:
                # Название занято чем-то посторонним — не тем же циклом: имя
                # несёт дату, и наше собственное совпадение уже поймала бы
                # проверка выше по этому же циклу. Останавливаем это правило,
                # остальные шаблоны идут дальше своим чередом.
                сбой = errors
                break
            задачи_кэш.append(задача["path"].stem)
            copy_template_attachments(имя, задача["path"].stem, today)
            созданы.append({"date": решение["date"].isoformat(), "task": задача["path"].stem})
            запись["previous"] = {"date": решение["date"].isoformat(), "closed": False,
                                  "task": задача["path"].stem}

        if созданы:
            state[имя] = запись
        if сбой:
            отчёт.append({"template": имя, "errors": сбой,
                          "created": созданы, "skipped": пропущены})
        else:
            отчёт.append({"template": имя, "created": созданы, "skipped": пропущены})

    save_recurrence_state(state)
    return {"today": today.isoformat(), "templates": отчёт,
            "created": sum(len(t["created"]) for t in отчёт)}


def cmd_template_from_task(args, today):
    """Сделать шаблон из существующей задачи — «я это уже делал, повтори так же».

    Сдвиги считаются от даты первого шага, поэтому шаблон переносим на любую дату
    старта. Задача при этом не меняется.
    """
    задача = find_task(args.task)
    склад = tpl.JsonStore(VAULT)
    try:
        шаблон = tpl.template_from_task(задача["meta"], name=args.name)
        склад.save(шаблон)
    except tpl.TemplateError as e:
        return {"ok": False, "errors": getattr(e, "errors", [{"field": None, "error": str(e)}])}
    return {"ok": True, "template": шаблон["name"], "from_task": задача["path"].stem,
            "steps": len(шаблон.get("steps") or [])}


def cmd_refresh(args, today):
    """Пересчитать сводку во всех задачах.

    Статус устаревает сам по себе: задача становится просроченной оттого, что
    прошёл день, а не оттого, что кто-то её трогал. Гонять перед утренней
    сборкой.

    Раньше второй прогон в тот же день не трогал ни одного файла: сводка
    сравнивалась с тем, что уже лежало на диске, — экономило запись и не
    заставляло Obsidian переиндексировать вольт впустую. В БД сравнивать не с
    чем: сводка нигде не хранится (колонок под неё нет, source of truth в
    шагах), поэтому каждый refresh честно пересчитывает и отдаёт всё заново, а
    не только «изменившееся». `--force` остался в контракте ответа, но
    поведение больше не меняет — запись каждый раз одна и та же дешёвая
    операция.
    """
    touched = []
    for task in load_tasks():
        save(task, today)
        touched.append({"task": task["path"].stem,
                        "status": task["meta"]["status"], "changed": True})
    return {"today": today.isoformat(), "written": touched, "count": len(touched),
            "forced": bool(args.force), "broken": list(BROKEN)}


def parse_period_date(text, today):
    """Дата для границы периода — «с даты», «по дату» в архиве.

    Отдельно от `parse_date_input` намеренно, а не тот же вызов: тот планирует
    вперёд («пн» значит ближайший понедельник **после** сегодня — так и должен
    вести себя срок задачи), а граница периода смотрит в прошлое. «18.08» без
    года при разборе вперёд ушло бы в следующий август — для фильтра истории
    это означало бы «ничего не найдено» вместо прошлого августа, который
    заказчик и имел в виду. Здесь бортик наоборот: день без года берётся не
    позже сегодня, а если такой день ещё не наступил в этом году — годом раньше.
    """
    if text is None:
        return None
    s = str(text).strip().lower().replace("ё", "е")
    if not s:
        return None
    if s == "сегодня":
        return today
    if s == "вчера":
        return today - timedelta(days=1)
    if s == "позавчера":
        return today - timedelta(days=2)
    m = re.fullmatch(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    m = re.fullmatch(r"(\d{1,2})[.\-/](\d{1,2})(?:[.\-/](\d{2,4}))?", s)
    if m:
        день, месяц = int(m.group(1)), int(m.group(2))
        if m.group(3):
            год = int(m.group(3))
            return date(год + 2000 if год < 100 else год, месяц, день)
        try:
            кандидат = date(today.year, месяц, день)
        except ValueError:
            raise ValueError(f"не дата: {text!r}")
        return кандидат if кандидат <= today else date(today.year - 1, месяц, день)
    raise ValueError(f"не дата: {text!r}")


def _finished_date(task):
    """Дата, по которой архив фильтрует задачу периодом.

    Для закрытой — самое позднее `completed_date` среди листьев: момент, когда
    задача реально завершилась, а не когда её завели. Для отменённой листья
    даты закрытия могут не нести вовсе (отмена не требует закрывать шаги), тогда
    берётся дата начала — точнее взять неоткуда, а вовсе не фильтровать её по
    периоду означало бы, что отменённые задачи не находятся периодом никогда.
    """
    листья = leaves_of(task)
    завершения = [as_date(s["completed_date"]) for s in листья if s.get("completed_date")]
    if завершения:
        return max(завершения)
    return as_date(task["meta"].get("start_date"))


def cmd_archive(args, today):
    """История: закрытые и отменённые задачи. Раздел 5.9 ТЗ, требование R24.

    Фильтры — тег и период, дальше по тексту ищет `cmd_search`: полнотекстовый
    поиск и просмотр списком — разные операции с разными ответами, а не одна
    команда с необязательными полями.

    Циклы одного повторения сворачиваются в одну строку с возможностью
    развернуть (решение по Q24: заказчик выбрал именно так, а не список всех
    экземпляров подряд). Свёртка — по `template_name`, тому самому полю из
    схемы v3: группа получает `kind="cycle_group"` и список задач внутри,
    одиночная задача или шаблон с единственным закрытым циклом — `kind="task"`,
    сворачивать нечего.
    """
    тег = getattr(args, "tag", None)
    ошибки = []
    since = until = None
    try:
        if getattr(args, "since", None):
            since = parse_period_date(args.since, today)
    except ValueError:
        ошибки.append({"field": "since", "error": f"Дату не понял: «{args.since}»"})
    try:
        if getattr(args, "until", None):
            until = parse_period_date(args.until, today)
    except ValueError:
        ошибки.append({"field": "until", "error": f"Дату не понял: «{args.until}»"})
    if ошибки:
        return {"ok": False, "errors": ошибки}

    строки = []
    for task in load_tasks():
        status = task_status(task, today)
        if status not in ("done", "cancelled"):
            continue
        теги = task["meta"].get("tags") or []
        if тег and тег not in теги:
            continue
        когда = _finished_date(task)
        if since and (когда is None or когда < since):
            continue
        if until and (когда is None or когда > until):
            continue
        листья = leaves_of(task)
        строки.append({
            "task": task["path"].stem,
            "status": STATUS_RU[status],
            "date": когда.isoformat() if когда else None,
            "category": теги,
            "steps_total": len(листья),
            "template_name": task["meta"].get("template_name"),
            "cycle_key": task["meta"].get("cycle_key"),
        })

    группы, одиночные = {}, []
    for r in строки:
        (группы.setdefault(r["template_name"], []) if r["template_name"] else одиночные).append(r)

    out = []
    for имя, циклы in группы.items():
        циклы.sort(key=lambda r: r["cycle_key"] or "", reverse=True)
        if len(циклы) > 1:
            out.append({"kind": "cycle_group", "template_name": имя,
                       "count": len(циклы), "tasks": циклы})
        else:
            out.append({"kind": "task", **циклы[0]})
    for r in одиночные:
        out.append({"kind": "task", **r})

    def _дата_сортировки(item):
        if item["kind"] == "cycle_group":
            даты = [t["date"] for t in item["tasks"] if t["date"]]
            return max(даты) if даты else ""
        return item.get("date") or ""

    out.sort(key=_дата_сортировки, reverse=True)
    return {"ok": True, "today": today.isoformat(), "count": len(строки), "items": out}


def cmd_list(args, today):
    out = []
    for task in load_tasks():
        листья = leaves_of(task)
        step = current_step(task)
        out.append({
            "task": task["path"].stem,
            "status": task_status(task, today),
            "category": task["meta"].get("tags") or [],
            "steps_done": sum(1 for s in листья if s.get("status") in (DONE, SKIPPED)),
            "steps_total": len(листья),
            "current": step.get("title") if step else None,
        })
    return {"today": today.isoformat(), "tasks": out, "broken": list(BROKEN)}


def cmd_show(args, today):
    """Задача целиком — для карточки. `state` шага (просрочен/сегодня/ждёт)
    считается тут же, а не на странице: карточка не знает про рабочее время
    и не должна вычислять просрочку сама, тот же принцип, что у ленты.

    Заметка отдаётся без блока шагов: блок пишет `render_steps`, и в поле
    карточки ему делать нечего. Без `body` в ответе карточка показывала пустую
    заметку и отправляла эту пустоту обратно при первом же сохранении —
    заметка заказчика стиралась, хотя он её не трогал.

    Сводка (status/current_step/control_date/progress/stalled) в БД не хранится
    (см. save()), значит `task["meta"]` из Store.load_tasks() её не несёт —
    домешиваем `task_summary()` тем же приёмом, что и save(), иначе карточка
    получала бы задачу без срока и прогресса.
    """
    task = find_task(args.task)
    now = _now(args, today)
    work = _work(args)
    заметка, _ = _strip_steps_block(task["body"])
    meta = {**task["meta"], **task_summary(task, today)}
    закрыт, _дети = _closure(steps_of(task))
    return {
        "task": task["path"].stem,
        "status": task_status(task, today),   # английский, для сравнений в JS
        "body": заметка.strip(),
        "meta": {k: v for k, v in meta.items() if k != "steps"},  # meta["status"] — русский
        "steps": [
            {**{k: (str(v) if isinstance(v, (date, datetime)) else v)
                for k, v in s.items() if k != "log"},
             "log": s.get("log") or [],
             # closed отдаётся явно: у группы нет статуса, её закрытие
             # вычисляется из детей, и карточка не должна считать это сама
             "closed": закрыт(s),
             # stalled — по той же причине. Карточка считала его сама и считала
             # иначе, чем движок: брала `not_done` и `defer`, тогда как обычный
             # одиночный `defer` в счётчик намеренно не идёт (`stall_count`,
             # тест `test_defer_does_not_count_as_stalling`), зато идёт массовый
             # `mass_defer`. Один и тот же шаг показывал в ленте и в карточке
             # разные числа.
             "stalled": stall_count(s),
             "state": (None if is_group(s) else
                       worktime.due_state(s.get("control_date"), now, work)
                       if not is_closed(s) else None)}
            for s in steps_of(task)
        ],
    }


def cmd_done(args, today):
    task = find_task(args.task)
    step = get_step(task, args.step)
    if step.get("status") != OPEN:
        sys.exit(f"шаг {args.step} уже {step.get('status')}")
    step["status"] = DONE
    step["completed_date"] = today
    log_event(step, "done", today, reason=args.reason)

    # шаг без даты никогда не всплывёт в сборке — ставим сегодня, заказчик
    # увидит его и при необходимости перенесёт. Открыться могла параллельная
    # группа, то есть листьев несколько — дату получает каждый из них.
    активные = current_steps(task)
    assigned = [s["id"] for s in активные if not s.get("control_date")]
    for s in активные:
        if not s.get("control_date"):
            s["control_date"] = today
    save(task, today)
    return {"ok": True, "task": task["path"].stem, "step": args.step, "status": DONE,
            "next_step": активные[0].get("title") if активные else None,
            "date_assigned_to_step": assigned[0] if assigned else None,
            "dates_assigned": assigned,
            "task_status": task_status(task, today)}


def _reason_error(reason):
    """Причина обязана быть словом из справочника — раздел 5.4 ТЗ прямо
    говорит «из справочника», не «любой текст». Сравнение без регистра, тем же
    приёмом, что у тегов и шаблонов: «Не было денег» и «не было денег» — одна
    причина, а не две.

    Пустую причину сюда не пускают вызывающие: обязательна она или нет,
    решает конкретная операция (у провала — всегда, у переноса — оболочка),
    а не этот справочник. Здесь только «если указана — должна быть настоящей».
    """
    if not reason:
        return None
    if reason.strip().lower() not in {r.lower() for r in get_reasons()}:
        return {"field": "reason", "error": f"Причина «{reason}» не из справочника"}
    return None


def cmd_notdone(args, today):
    """Шаг остаётся открытым: причина обычно внешняя, решать всё равно надо."""
    ошибка = _reason_error(args.reason)
    if ошибка:
        return {"ok": False, "errors": [ошибка]}
    task = find_task(args.task)
    step = get_step(task, args.step)
    if step.get("status") != OPEN:
        sys.exit(f"шаг {args.step} уже {step.get('status')}")
    new_date = parse_stored_control(args.to) if args.to else today + timedelta(days=1)
    log_event(step, "not_done", today, reason=args.reason,
              was=as_date(step.get("control_date")), to=new_date)
    step["control_date"] = new_date
    save(task, today)
    count = stall_count(step)
    return {"ok": True, "task": task["path"].stem, "step": args.step, "status": OPEN,
            "next_check": tpl.control_text(new_date), "stalled": count,
            "hint": "шаг буксует, нужен другой ход" if count >= 3 else None}


def _defer_step(task, step, today, to, reason, event="defer"):
    """Общая механика переноса: пишет событие в журнал шага, двигает дату,
    сохраняет задачу. Используется одиночным `defer` и массовым переносом из
    разбора завала (R20 ТЗ) — второй передаёт `event="mass_defer"`, чтобы
    запись в журнале была отличима от обычного переноса."""
    log_event(step, event, today, reason=reason,
              was=as_date(step.get("control_date")), to=to)
    step["control_date"] = to
    save(task, today)


def cmd_defer(args, today):
    ошибка = _reason_error(args.reason)
    if ошибка:
        return {"ok": False, "errors": [ошибка]}
    task = find_task(args.task)
    step = get_step(task, args.step)
    if step.get("status") != OPEN:
        sys.exit(f"шаг {args.step} уже {step.get('status')}")
    to = parse_stored_control(args.to)
    _defer_step(task, step, today, to, args.reason)
    return {"ok": True, "task": task["path"].stem, "step": args.step,
            "next_check": tpl.control_text(to)}


def cmd_fail(args, today):
    """«Не будет сделано» — шаг закрывается проваленным, задача идёт дальше.

    Четвёртый исход из раздела 6.4 ТЗ, намеренно менее заметный в интерфейсе:
    он не должен становиться лёгким путём отмахнуться. Отличается от «снят» тем,
    что снятый шаг перестал быть нужен, а проваленный был нужен и не случился —
    и в истории это разные вещи.
    """
    ошибка = _reason_error(args.reason)
    if ошибка:
        return {"ok": False, "errors": [ошибка]}
    task = find_task(args.task)
    step = get_step(task, args.step)
    if is_closed(step):
        sys.exit(f"шаг {args.step} уже {step.get('status')}")
    step["status"] = FAILED
    log_event(step, "failed", today, reason=args.reason)
    активные = current_steps(task)
    for s in активные:
        if not s.get("control_date"):
            s["control_date"] = today
    save(task, today)
    return {"ok": True, "task": task["path"].stem, "step": args.step, "status": FAILED,
            "next_step": активные[0].get("id") if активные else None,
            "task_status": task_status(task, today)}


def cmd_skip(args, today):
    """Шаг снят: задача пошла другим путём, а не через этот шаг."""
    task = find_task(args.task)
    step = get_step(task, args.step)
    # Как и done/notdone/defer: закрытый шаг повторно не трогаем. Иначе снятие
    # уже сделанного шага оставляло бы completed_date и событие «сделан» в логе
    # рядом со статусом «снят» — запись, противоречащая сама себе.
    if is_closed(step):
        sys.exit(f"шаг {args.step} уже {step.get('status')}")
    step["status"] = SKIPPED
    log_event(step, "skipped", today, reason=args.reason)
    for s in current_steps(task):
        if not s.get("control_date"):
            s["control_date"] = today
    save(task, today)
    return {"ok": True, "task": task["path"].stem, "step": args.step, "status": SKIPPED,
            "task_status": task_status(task, today)}


def cmd_backlog_bulk(args, today):
    """Массовые действия из разбора завала — вкладка «Списком», R20 ТЗ.

    Три операции над списком `{task, step}`: перенос всей пачки на одну дату
    с одной причиной, массовое «сделано», массовое «не будет сделано». Каждый
    элемент проходит ту же механику, что и одиночные `defer`/`done`/`fail` —
    здесь ничего не дублируется, элементы `done`/`fail` зовут сами эти команды,
    перенос зовёт общий с `cmd_defer` `_defer_step`.

    Один плохой элемент не роняет пачку: вольт правится руками и другой
    процесс мог закрыть тот же шаг за это время, поэтому каждый элемент — свой
    try/except, а не общий. По каждому отдаётся успех или структурная ошибка
    {field, error} — то же правило, что и у одиночных операций (КОНТРАКТ.md).
    Повтор пачки после обрыва связи не портит уже закрытые шаги: они просто
    попадут в ответе как ошибка «уже done/failed», а не продублируют запись —
    та же идемпотентность, что у одиночных команд.

    Причина обязательна для переноса и «не будет сделано», как и у одиночных
    `defer`/`fail`; для «сделано» — нет, как и у одиночного `done`. Дата для
    переноса приходит уже разобранной (ISO) — человеческий ввод вроде «+3»
    разбирает `parse_date_input`, а не эта команда, и не оболочка.
    """
    op = args.op
    items = args.items or []
    if isinstance(items, str):
        # CLI отдаёт JSON-строку (или "-" для stdin), сервер — уже готовый
        # список: тот же приём, что у `cmd_create` с телом задачи.
        сырое = sys.stdin.read() if items == "-" else items
        try:
            items = json.loads(сырое)
        except json.JSONDecodeError as e:
            return {"ok": False, "errors": [{"field": "items", "error": f"битый JSON: {e}"}]}
    reason = (args.reason or "").strip() or None
    to = parse_stored_control(args.to) if getattr(args, "to", None) else None

    if op not in ("defer", "done", "fail"):
        return {"ok": False, "errors": [{"field": "op", "error": f"неизвестное действие: {op}"}]}
    if op in ("defer", "fail") and not reason:
        return {"ok": False, "errors": [{"field": "reason", "error": "причина обязательна"}]}
    ошибка = _reason_error(reason)
    if ошибка:
        return {"ok": False, "errors": [ошибка]}
    if op == "defer" and to is None:
        return {"ok": False, "errors": [{"field": "to", "error": "нужна новая дата"}]}

    results = []
    for позиция in items:
        позиция = позиция or {}
        имя_задачи = позиция.get("task")
        id_шага = позиция.get("step")
        строка_шага = str(id_шага) if id_шага not in (None, "") else None
        try:
            if not имя_задачи or id_шага in (None, ""):
                raise ValueError("не указан шаг")
            if op == "defer":
                task = find_task(имя_задачи)
                step = get_step(task, id_шага)
                if step.get("status") != OPEN:
                    raise ValueError(f"шаг {id_шага} уже {step.get('status')}")
                _defer_step(task, step, today, to, reason, event="mass_defer")
                результат = {"ok": True, "task": task["path"].stem, "step": строка_шага,
                            "next_check": tpl.control_text(to), "stalled": stall_count(step)}
            elif op == "done":
                результат = cmd_done(
                    SimpleNamespace(task=имя_задачи, step=строка_шага, reason=reason), today)
            else:  # fail
                результат = cmd_fail(
                    SimpleNamespace(task=имя_задачи, step=строка_шага, reason=reason), today)
            results.append(результат)
        except (SystemExit, ValueError) as e:
            results.append({"ok": False, "task": имя_задачи, "step": строка_шага,
                            "errors": [{"field": None, "error": str(e)}]})

    return {"op": op, "count": len(results),
            "ok_count": sum(1 for r in results if r["ok"]),
            "fail_count": sum(1 for r in results if not r["ok"]),
            "items": results}


# --- база знаний -------------------------------------------------------

def _read_json_arg(value):
    """JSON: строкой, уже разобранным значением (так приходит из HTTP — сервер
    парсит тело запроса раньше) или «-» для чтения из stdin, как у `create`.
    Одна команда обслуживает и форму, и командную строку, и вход у них разный."""
    if isinstance(value, str):
        if value == "-":
            value = sys.stdin.read()
        return json.loads(value)
    return value


def _kb_stores():
    """Склады ссылок и отказов: БД, если база знаний туда переехала, иначе файлы.

    Признак тот же, что у `load_kb_entries`, — есть ли записи в `kb_notes`.
    Держать его в одном месте обязательно: разъехавшись, чтение записей и
    чтение ссылок к ним начнут смотреть в разные хранилища, и подтверждённое
    подчёркивание перестанет находиться.
    """
    if get_store().load_kb_notes():
        return kb.SqliteLinkStore(VAULT), kb.SqliteExclusionStore(VAULT)
    return kb.JsonLinkStore(VAULT), kb.JsonExclusionStore(VAULT)


def migrate_kb_to_db(today=None):
    """Перенести базу знаний из markdown в таблицы. Этап (b) плана.

    Что переносится: `База/*.md` → `kb_notes`, `Ссылки.json` → `kb_links`,
    `Исключения.json` → `kb_exclusions`. Файлы после переноса остаются на диске
    нетронутыми: это данные заказчика, и удалять их за него мы не будем — но
    источником правды они быть перестают (см. `load_kb_entries`).

    Главная работа тут не в переливании, а в **смене идентификатора**. В
    markdown записью правило имя файла, то есть строка «Василий Говнов», и
    ссылки в `Ссылки.json` ссылаются ею. В базе id числовой. Поэтому запись
    сохраняет своё прежнее имя в `legacy_file`, а ссылки перецепляются по нему.
    Ссылка на запись, которой в `База/` уже нет (файл удалили руками, а ссылка
    осталась), переносу не подлежит: в базе внешний ключ, и висячая строка туда
    просто не ляжет. Такие считаются и возвращаются числом, а не выбрасываются
    молча — заказчику стоит знать, что часть подчёркиваний исчезнет.

    Идемпотентно: если в `kb_notes` уже что-то есть, второй прогон ничего не
    делает. Повторный запуск после обрыва не должен заводить вторые копии.
    """
    склад = get_store()
    if склад.load_kb_notes():
        return {"ok": True, "skipped": "база знаний уже в БД",
                "notes": 0, "links": 0, "exclusions": 0, "dropped_links": 0}

    записи = load_kb_entries()   # здесь ещё markdown: таблица пуста
    имя_к_id = {}
    for з in записи:
        тело = ""
        путь = KB_DIR / f"{з['id']}.md"
        if путь.is_file():
            try:
                _, тело = parse_file(путь)
            except Exception:
                тело = ""
        имя_к_id[str(з["id"])] = склад.add_kb_note(
            з["title"], з.get("aliases") or [], тело, legacy_file=str(з["id"]))

    старые_ссылки = kb.JsonLinkStore(VAULT).all() if (VAULT / "Ссылки.json").is_file() else []
    новые, потеряно = [], 0
    for с in старые_ссылки:
        новый_id = имя_к_id.get(str(с.get("kb_entry_id")))
        if новый_id is None:
            потеряно += 1
            continue
        новые.append({**с, "kb_entry_id": новый_id})
    склад.save_kb_links(новые)

    старые_отказы = (kb.JsonExclusionStore(VAULT).keys()
                     if (VAULT / "Исключения.json").is_file() else set())
    отказы = []
    for запись, написание in старые_отказы:
        # None в первом поле — «слово никогда не ссылка, у любой записи»: такой
        # отказ не привязан к записи и переезжает как есть.
        отказы.append({"kb_entry_id": имя_к_id.get(str(запись)) if запись is not None else None,
                       "text": написание})
    склад.save_kb_exclusions([о for о in отказы
                              if о["kb_entry_id"] is not None or о["text"]])

    return {"ok": True, "notes": len(имя_к_id), "links": len(новые),
            "exclusions": len(отказы), "dropped_links": потеряно}


def cmd_migrate_kb(args, today):
    """Команда для этапа (b). Отдельная и запускаемая руками, а не при старте:
    миграция трогает данные, и делать это молча в фоне нельзя."""
    return migrate_kb_to_db(today)


def _search_text_of_task(task):
    """Что от задачи попадает в поиск: заголовок, заметка и названия шагов.

    Причины переносов сюда не идут — решение по Q21 («пока не нужно»). Когда
    понадобятся, они станут отдельными строками индекса со своим `source_type`,
    и ни эта функция, ни форма таблицы не изменятся.
    """
    куски = [task["path"].stem, (task.get("body") or "")]
    куски += [s.get("title") or "" for s in steps_of(task)]
    return "\n".join(к for к in куски if к)


def reindex_search(store_=None):
    """Собрать поисковый индекс заново по всему вольту. Требования R24, R25.

    Полная пересборка, а не досборка: она нужна после переезда, после смены
    правил лемматизации и как способ починить индекс, если он разошёлся с
    данными. На целевом объёме ТЗ это секунды, а разошедшийся индекс чинится
    иначе только руками.

    Дальше индекс поддерживается по одной задаче в `save()` — там своя строка
    переписывается, а не пересобирается всё.
    """
    склад = store_ or get_store()
    склад.search_clear()
    задач = 0
    for task in load_tasks():
        склад.search_replace(
            "task", task["path"].stem, task["path"].stem,
            (task.get("body") or "").strip()[:200],
            kb.lemmatize_text(_search_text_of_task(task)))
        задач += 1
    записей = 0
    for з in склад.load_kb_notes():
        склад.search_replace(
            "kb_note", з["id"], з["title"], (з.get("body") or "").strip()[:200],
            kb.lemmatize_text(" ".join([з["title"], *(з.get("aliases") or []),
                                        з.get("body") or ""])))
        записей += 1
    return {"ok": True, "tasks": задач, "kb_notes": записей}


def cmd_reindex(args, today):
    return reindex_search()


def search_query(text):
    """Человеческий запрос → выражение FTS5.

    Слова приводятся к тем же леммам, что и текст в индексе: иначе «гранту» в
    поиске не нашло бы «грант» в задаче, ради чего лемматизация и заводилась.
    Слова соединяются через AND — человек, набравший два слова, ищет то, где
    есть оба, а не то, где есть хоть одно.

    Каждое слово берётся в кавычки: в запрос попадают знаки, которые FTS5
    считает синтаксисом (дефис в «финмодель-2026», звёздочка, скобки), и без
    кавычек он на них ругается или понимает их не так, как человек имел в виду.
    """
    слова = [w for w, _, _ in kb.tokenize(text or "")]
    if not слова:
        return ""
    return " AND ".join(f'"{kb.lemma(w)}"' for w in слова)


def cmd_search(args, today):
    """Поиск по истории — R24, главный ответ на «как я это делал в прошлый раз».

    Ищет по задачам, их заметкам, названиям шагов и записям базы знаний.
    """
    запрос = search_query(getattr(args, "text", None))
    if not запрос:
        return {"ok": False, "errors": [{"field": "text", "error": "Пустой запрос"}]}
    виды = None
    if getattr(args, "kind", None):
        виды = [args.kind]
    найдено = get_store().search(запрос, limit=int(getattr(args, "limit", None) or 50),
                                 source_types=виды)
    return {"ok": True, "query": запрос, "count": len(найдено), "results": найдено}


def _kb_settings():
    """Настройки автораспознавания. Тот же приём, что `_work`/`_backup_settings`:
    файл — база, битый файл не роняет сканирование."""
    try:
        сохранённые = cfg.load(cfg.settings_path(VAULT))["kb"]
    except cfg.SettingsError:
        сохранённые = cfg.defaults()["kb"]
    return сохранённые


def cmd_kb_scan(args, today):
    """Гипотезы упоминаний записей базы знаний в тексте — R17, разделы 5.7/5.8 ТЗ.

    Индекс собирается заново на каждый вызов, не кэшируется: заказчик правит
    заметки базы в Obsidian, а кэш без инвалидации на эту правку не среагирует.

    `kb.auto_recognition` выключает поиск новых совпадений целиком — заказчик
    решил, что подчёркивания мешают, а не сканирование сломано. Уже
    подтверждённые ссылки при этом продолжают показываться: это не гипотезы,
    а факт, который заказчик когда-то подтвердил сам, и выключенный флаг не
    должен стирать историю. `kb.min_match_length` идёт в `build_index` вместо
    зашитой в `kb.py` константы — раньше это поле лежало в файле настроек,
    а искало ровно четыре буквы, что бы там ни было записано.
    """
    настройки = _kb_settings()
    source_type = getattr(args, "source_type", None)
    source_id = getattr(args, "source_id", None)
    _ссылки_склад, _отказы_склад = _kb_stores()
    подтверждённые = (_ссылки_склад.for_source(source_type, source_id)
                      if source_type and source_id else [])

    text = getattr(args, "text", None) or ""
    entries = load_kb_entries()
    if not text or not entries or not настройки.get("auto_recognition", True):
        return {"hypotheses": [], "confirmed": подтверждённые, "kb_broken": list(KB_BROKEN)}

    исключения = _отказы_склад.keys()
    индекс = kb.build_index(entries, min_match=настройки.get("min_match_length") or kb.MIN_MATCH)
    гипотезы = kb.find_mentions(text, индекс, excluded=исключения)

    if подтверждённые:
        # Уже отвеченное не переспрашиваем: смещения подтверждённых ссылок
        # исключаются из новых гипотез по тому же месту в тексте.
        занято = {(с["offset_start"], с["offset_end"]) for с in подтверждённые}
        гипотезы = [г for г in гипотезы
                    if (г["offset_start"], г["offset_end"]) not in занято]

    return {"hypotheses": гипотезы, "confirmed": подтверждённые,
            "kb_broken": list(KB_BROKEN)}


def _find_task_by_stem(name):
    """Точное имя файла задачи, без нечёткого поиска `find_task`: гипотезы уже
    посчитаны на конкретной, уже созданной задаче — мазать мимо здесь нельзя."""
    for t in load_tasks():
        if t["path"].stem == name:
            return t
    return None


def _strip_steps_block(body):
    """Тело без блока шагов плюс способ вернуть блок на прежнее место.

    Смещения гипотез посчитаны против текста БЕЗ этого блока (см. `cmd_kb_scan`
    и докстринг вызывающего кода), поэтому сплайс ссылок должен идти по той же
    системе координат. Блока может не быть вовсе — задача только что создана
    и ещё не проходила через `save()`; тогда возвращается тело как есть.

    Первая же вставка блока (`put_steps_into_body`, когда маркеров ещё не
    было) склеивает его с текстом заказчика через «\\n\\n» — до одного
    перевода строки, если текста не было вовсе. Эта склейка не текст
    заказчика, а механика рендера, и в форме создания на момент сканирования
    её ещё нет. Не срезать её здесь — значит увести смещения на два символа
    для любой задачи, которую подтверждают сразу после первого сохранения:
    ровно тот путь, которым и приходит подтверждение из формы (раздел 4).
    Дальше эта склейка не меняется (`put_steps_into_body` при найденных
    маркерах переносит хвост как есть), поэтому срез безопасен и на
    повторных сохранениях — режется всегда один и тот же кусок.
    """
    start = body.find(STEPS_START)
    end = body.find(STEPS_END)
    if start == -1 or end == -1 or end <= start:
        return body, lambda stripped: stripped
    block_end = end + len(STEPS_END)
    block = body[start:block_end]
    tail = body[block_end:]
    склейка = tail[:2] if tail[:2] == "\n\n" else (tail[:1] if tail[:1] == "\n" else "")
    return (body[:start] + tail[len(склейка):],
            lambda stripped: stripped[:start] + block + склейка + stripped[start:])


def _mark_links_in_body(task, mentions):
    """Вписать подтверждённые упоминания в тело настоящими вики-ссылками.

    Единственное место, где текст заказчика правит программа, — и делает это
    ровно потому, что человек сам нажал «да» на конкретное упоминание (раздел
    7.1 ТЗ). Пишет `[[Название]]`, а если написано не так, как называется
    запись — piped link `[[Название|как написано]]`.

    Гипотезы обрабатываются по убыванию offset_start: иначе первая же вставка
    сдвинет смещения соседних. Та, под которой текст успел измениться
    (`text[s:e] != matched`), пропускается молча — ссылка в Ссылки.json уже
    записана вызывающим, здесь только подчёркивание в тексте.
    """
    body, вернуть_блок = _strip_steps_block(task["body"])
    for гипотеза in sorted(mentions, key=lambda г: -г["offset_start"]):
        s, e = гипотеза["offset_start"], гипотеза["offset_end"]
        if body[s:e] != гипотеза["matched"]:
            continue
        title, matched = гипотеза["title"], гипотеза["matched"]
        link = f"[[{title}]]" if matched == title else f"[[{title}|{matched}]]"
        body = body[:s] + link + body[e:]
    task["body"] = вернуть_блок(body)


def cmd_kb_confirm(args, today):
    """Подтверждение гипотез автораспознавания — R17, раздел 7.1 ТЗ.

    Каждая гипотеза обрабатывается независимо: одна кривая не блокирует
    соседние. Ошибки собираются с индексом гипотезы в поле (`mentions.0` и
    т. п.), `ok` — False только если упали все.
    """
    source_type = getattr(args, "source_type", None)
    source_id = getattr(args, "source_id", None)
    try:
        mentions = _read_json_arg(args.mentions) or []
    except json.JSONDecodeError as e:
        return {"ok": False, "links": [],
                "errors": [{"field": "mentions", "error": f"битый JSON: {e}"}]}

    склад, _ = _kb_stores()
    успешные, ошибки = [], []
    for i, гипотеза in enumerate(mentions):
        try:
            ссылка = склад.add(гипотеза, source_type=source_type, source_id=source_id)
        except kb.KbError as e:
            for err in e.errors:
                поле = f"mentions.{i}.{err['field']}" if err.get("field") else f"mentions.{i}"
                ошибки.append({"field": поле, "error": err["error"]})
            continue
        успешные.append((гипотеза, ссылка))

    if успешные and source_type == "task":
        задача = _find_task_by_stem(source_id)
        if задача is not None:
            _mark_links_in_body(задача, [г for г, _ in успешные])
            save(задача, today)

    ok = bool(успешные) or not ошибки
    return {"ok": ok, "links": [с for _, с in успешные], "errors": ошибки}


def cmd_kb_reject(args, today):
    """Ответ «нет» на гипотезу — раздел 7.1 ТЗ.

    `mute=True` — слово никогда не считается ссылкой ни у одной записи
    («Грант» у заказчика чаще сумма денег, чем запись базы). `mute=False` —
    отказ только от этой конкретной гипотезы у этой записи.
    """
    try:
        гипотеза = _read_json_arg(args.mention)
    except json.JSONDecodeError as e:
        return {"ok": False, "errors": [{"field": "mention", "error": f"битый JSON: {e}"}]}

    _, склад = _kb_stores()
    if getattr(args, "mute", False):
        склад.mute_word(гипотеза["matched"])
    else:
        склад.reject(гипотеза)
    return {"ok": True}


# --- выгрузка в Excel ------------------------------------------------------

# Предел Excel на длину текста в ячейке. Тело заметки пишет заказчик, и упереться
# в него теоретически можно — лучше обрезать, чем получить нечитаемый файл.
MAX_CELL = 32767

# Управляющие символы xlsx не принимает: файл открывается с руганью на повреждение.
# Тело заметки приходит из внешнего редактора, так что чистим на всякий случай.
CONTROL_CHARS = re.compile(r"[\000-\010\013\014\016-\037]")

# Заголовок и ширина колонки. Ширину задаём руками, а не по содержимому: имена
# задач и причины переносов длинные, и по факту всё равно упираешься в потолок,
# а файл должен открываться готовым к чтению, без растаскивания колонок мышью.
TASK_COLUMNS = [
    ("Задача", 34), ("Заголовок", 30), ("Создана", 12), ("Статус", 14),
    ("Текущий шаг", 34), ("Контроль", 12), ("Прогресс", 10), ("Буксует", 9),
    ("Категории", 22), ("Заметка", 60),
]
STEP_COLUMNS = [
    ("Задача", 34), ("Шаг", 6), ("Название", 40), ("Вид", 10), ("В группе", 30),
    ("Статус", 10), ("Контроль", 12), ("Выполнен", 12), ("Не сделан, раз", 15),
    ("Последняя причина", 40),
]
EVENT_COLUMNS = [
    ("Задача", 34), ("Шаг", 6), ("Название", 34), ("Дата", 12),
    ("Событие", 14), ("Причина", 40), ("Было", 12), ("Стало", 12),
]


def put(ws, row, col, value):
    """Одна ячейка со всеми оговорками про Excel.

    Даты кладём объектами `date`: строкой Excel их не понимает, теряется сортировка
    и фильтр по периоду — та же причина, по которой даты пишутся датами и в YAML.

    Строку, начинающуюся с «=», openpyxl считает формулой. В теле заметки такая
    строка вполне возможна, и Excel потом ругается на весь файл — тип задаём явно.
    """
    cell = ws.cell(row=row, column=col)
    if isinstance(value, datetime):
        value = value.date()
    if value == "":
        value = None  # задача без категорий и без тела — просто пустая ячейка
    if isinstance(value, date):
        cell.value = value
        cell.number_format = "YYYY-MM-DD"
    elif isinstance(value, str):
        value = CONTROL_CHARS.sub("", value)
        if len(value) > MAX_CELL:
            value = value[:MAX_CELL - 1] + "…"
        cell.value = value
        cell.data_type = "s"
    else:
        cell.value = value
    return cell


def write_sheet(wb, title, columns, rows):
    """Лист целиком: шапка, ширины, данные, фильтр."""
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet(title)
    for i, (name, width) in enumerate(columns, start=1):
        ws.cell(row=1, column=i, value=name).font = Font(bold=True)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"  # шапка не уезжает при прокрутке длинной истории
    for r, values in enumerate(rows, start=2):
        for c, value in enumerate(values, start=1):
            put(ws, r, c, value)
    # Фильтр по шапке: сортировка по дате и отбор по задаче — первое, что человек
    # захочет сделать в тысяче строк истории.
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{ws.max_row}"
    return ws


def cmd_export(args, today):
    """Весь вольт в один .xlsx.

    Заказчику это не аналитика, а страховка «чтобы не проебалось»: файл уезжает ему
    в Telegram и лежит отдельно от вольта. Отсюда третий лист — плоский разворот
    `log` шагов. Сводку и статусы можно пересчитать из шагов заново, а переносы,
    причины и старые даты больше взять неоткуда: пропадут вместе с папкой.

    Только чтение: вольт после экспорта побайтово тот же.
    """
    # Импорт по требованию: openpyxl нужен одной команде из восьми, а грузится он
    # впятеро дольше всего остального движка. Отметку шага это тормозить не должно.
    try:
        from openpyxl import Workbook
    except ImportError:
        sys.exit("нужен openpyxl: pip install openpyxl")

    out = Path(args.to) if args.to else VAULT / f"Выгрузка {today.isoformat()}.xlsx"
    tasks, steps, events = [], [], []

    for task in load_tasks():
        name = task["path"].stem
        meta = task["meta"]
        # Статус и прогресс считаем заново, а не берём из файла: сводка устаревает
        # сама по себе, от того что прошёл день. В выгрузке должно стоять сегодня.
        status = task_status(task, today)
        step = current_step(task)
        all_steps = steps_of(task)
        по_id = {s["id"]: s for s in all_steps}
        листья = [s for s in all_steps if not is_group(s)]
        closed = sum(1 for s in листья if s.get("status") in (DONE, SKIPPED))
        tags = meta.get("tags") or []
        if isinstance(tags, str):
            tags = [tags]
        tasks.append([
            name, meta.get("title"), as_date(meta.get("created")), STATUS_RU[status],
            step.get("title") if step else None,
            as_date(step.get("control_date")) if step else None,
            f"{closed}/{len(листья)}" if листья else None,
            stall_count(step) if step else 0,
            ", ".join(str(t) for t in tags), task["body"].strip(),
        ])

        for s in all_steps:
            step_title = s.get("title")
            родитель = по_id.get(s.get("parent"))
            steps.append([
                name, s.get("id"), step_title,
                ("группа ∥" if s.get("mode") == "par" else
                 "группа →" if s.get("mode") == "seq" else None),
                родитель.get("title") if родитель else None,
                (None if is_group(s) else
                 STEP_STATUS_RU.get(s.get("status", OPEN), s.get("status"))),
                as_date(s.get("control_date")), as_date(s.get("completed_date")),
                stall_count(s),
                next((e.get("reason") for e in reversed(s.get("log") or [])
                      if e.get("reason")), None),
            ])
            for e in s.get("log") or []:
                events.append([
                    name, s.get("id"), step_title, as_date(e.get("date")),
                    EVENT_RU.get(e.get("event"), e.get("event")), e.get("reason"),
                    as_date(e.get("was")), as_date(e.get("to")),
                ])

    wb = Workbook()
    wb.remove(wb.active)  # лист по умолчанию называется Sheet и нам не нужен
    for name, columns, rows in (("Задачи", TASK_COLUMNS, tasks),
                                ("Шаги", STEP_COLUMNS, steps),
                                ("История", EVENT_COLUMNS, events)):
        write_sheet(wb, name, columns, rows)

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return {"today": today.isoformat(), "file": str(out), "tasks": len(tasks),
            "steps": len(steps), "events": len(events)}


# --- резервные копии и экспорт (R25, R11 — раздел 9 ТЗ) --------------------
#
# Копия (R25) и экспорт (R11) — разные задачи, backup.py разводит их по разным
# функциям. Здесь только тонкая обвязка: разбор аргументов, дефолтный путь и
# перевод backup.BackupError в структурную ошибку контракта {field, error}.

def default_backup_dir():
    """Куда класть копии по умолчанию — рядом с вольтом, не внутри него.

    Если пропадёт папка вольта целиком (отвалившийся диск, случайное
    удаление), копия обязана остаться цела. Папку заранее не создаём:
    `backup.create` заводит её сама при первом снятии копии.
    """
    return VAULT.parent / f"{VAULT.name} — копии"


def _backup_settings():
    """Папка, частота и число копий из настроек вольта.

    Читается тем же приёмом, что рабочие часы в `_work`: файл — база, битый
    файл не роняет операцию. Раньше эти три ключа лежали в `Настройки.json`
    мёртвыми: форма их не показывала, а `cmd_backup` брал папку рядом с вольтом
    и `backup.KEEP_DEFAULT`, что бы в них ни стояло.
    """
    try:
        сохранённые = cfg.load(cfg.settings_path(VAULT))["backup"]
    except cfg.SettingsError:
        сохранённые = cfg.defaults()["backup"]
    return сохранённые


def cmd_backup(args, today):
    """Снять резервную копию вольта. Раздел 9 ТЗ, требование R25."""
    настройки = _backup_settings()
    dest = Path(args.dest) if getattr(args, "dest", None) else (
        Path(настройки["folder"]) if настройки.get("folder") else default_backup_dir())
    keep = getattr(args, "keep", None) or настройки.get("keep_count") or backup.KEEP_DEFAULT
    try:
        итог = backup.backup(VAULT, dest, keep=keep, force=bool(getattr(args, "force", False)))
    except backup.BackupError as e:
        return {"ok": False, "errors": [e.as_json()]}
    except OSError as e:
        # Диск полон, папка недоступна — не наша ошибка с полем, но и молчать
        # нельзя: заказчик должен узнать, что копия не снялась.
        return {"ok": False, "errors": [{"field": None, "error": str(e)}]}
    return {"ok": True, **итог}


def cmd_backup_list(args, today):
    """Список копий для интерфейса восстановления. Пустого списка не боимся —
    первой копии могло ещё не быть, это не ошибка."""
    dest = Path(args.dest) if getattr(args, "dest", None) else default_backup_dir()
    return {"copies": backup.copies(dest), "dest": str(dest)}


def cmd_backup_restore(args, today):
    """Восстановить вольт из конкретной копии. Дефолта на «последнюю копию»
    нет намеренно: это деструктивная операция, файл выбирает человек.

    `backup.restore` сама снимает страховочную копию текущего состояния перед
    перезаписью — вызывающему снимать её отдельно не нужно.
    """
    try:
        итог = backup.restore(args.file, VAULT)
    except backup.BackupError as e:
        return {"ok": False, "errors": [e.as_json()]}
    return {"ok": True, **итог}


def cmd_export_json(args, today):
    """Выгрузка всей базы в JSON. Раздел 9 ТЗ, требование R11.

    Не Excel-выгрузка (`cmd_export`): та для «посмотреть глазами и переслать»,
    эта — данные без нашего формата хранения, для будущей версии продукта.

    Путь по умолчанию — рядом с копиями, не в самом вольте: та же логика, что
    у `default_backup_dir` — файл не должен пропасть вместе с папкой вольта.
    """
    out = (Path(args.to) if getattr(args, "to", None)
           else default_backup_dir() / f"выгрузка-{today.isoformat()}.json")
    try:
        итог = backup.write_export(VAULT, out)
    except backup.BackupError as e:
        return {"ok": False, "errors": [e.as_json()]}
    return {"ok": True, **итог}


# --- вложения ----------------------------------------------------------

def _attachment_owner(args):
    """(source_type, source_id) из `task`/`--step` или `--template`: та же
    адресация, что и везде в контракте — по названию, не по числовому id из
    базы. Задача обязана существовать; шаг, если указан, — тоже, иначе привязка
    ссылалась бы на то, чего нет.

    Шаблон адресуется именем из хранилища, а не тем, что прислали: имена
    сравниваются без регистра (`tpl.same_name`), и «квартальный отчёт» не
    должен завести вложениям вторую полку рядом с «Квартальный отчёт».
    """
    имя_шаблона = getattr(args, "template", None)
    if имя_шаблона not in (None, ""):
        шаблон = tpl.JsonStore(VAULT).get(имя_шаблона)
        if not шаблон:
            sys.exit(f"нет шаблона «{имя_шаблона}»")
        return "template", шаблон["name"]
    if getattr(args, "task", None) in (None, ""):
        sys.exit("нужно название задачи или --template")
    task = find_task(args.task)
    шаг = getattr(args, "step", None)
    if шаг not in (None, ""):
        try:
            шаг_id = int(шаг)
        except (TypeError, ValueError):
            шаг_id = None
        if шаг_id not in {s["id"] for s in steps_of(task)}:
            sys.exit(f"нет шага {шаг} в «{task['path'].stem}»")
        return "step", f'{task["path"].stem}:{шаг_id}'
    return "task", task["path"].stem


def cmd_attach(args, today):
    """Прикрепить файл к задаче или к шагу (`--step`). Один путь для CLI и
    формы: данные приходят либо уже готовыми байтами (форма — `args.data`),
    либо путём к файлу на диске (CLI — `args.file`) — тот же приём, что у
    `cmd_create` с JSON-строкой или `-` для чтения из stdin.
    """
    try:
        source_type, source_id = _attachment_owner(args)
    except SystemExit as e:
        поле = "template" if getattr(args, "template", None) else "task"
        return {"ok": False, "errors": [{"field": поле, "error": str(e)}]}

    filename = (getattr(args, "filename", None) or "").strip()
    if not filename and getattr(args, "file", None):
        filename = Path(args.file).name
    if not filename:
        return {"ok": False, "errors": [{"field": "filename", "error": "Нужно имя файла"}]}

    if getattr(args, "data", None) is not None:
        data = args.data
    else:
        try:
            data = Path(args.file).read_bytes()
        except OSError as e:
            return {"ok": False, "errors": [{"field": "file", "error": str(e)}]}

    try:
        sha256, size = attachments.save(VAULT, data, filename)
    except attachments.AttachmentError as e:
        return {"ok": False, "errors": [e.as_json()]}

    mime = attachments.guess_mime(filename)
    caption = (getattr(args, "caption", None) or "").strip() or None
    attachment_id = get_store().add_attachment(
        source_type, source_id, sha256, filename, mime, size, caption, today)
    return {"ok": True, "id": attachment_id, "sha256": sha256, "filename": filename,
            "mime": mime, "bytes": size}


def cmd_attachments(args, today):
    """Список вложений задачи целиком или одного шага (`--step`)."""
    source_type, source_id = _attachment_owner(args)
    rows = get_store().list_attachments(source_type, source_id)
    return {"attachments": [
        {"id": r["id"], "filename": r["filename"], "mime": r["mime"],
         "bytes": r["bytes"], "caption": r["caption"], "added": r["added"]}
        for r in rows]}


def cmd_attachment_delete(args, today):
    """Удалить вложение. Не задачу и не шаг — на связь между ними это никак
    не влияет, только на список вложений."""
    if get_store().get_attachment(args.id) is None:
        sys.exit(f"нет вложения {args.id}")
    get_store().delete_attachment(args.id)
    return {"ok": True, "id": args.id}


def rename_tag_everywhere(old_name, new_name, today=None):
    """Заменить тег во всех задачах вольта. Дополняет `settings.rename_tag`
    и `settings.merge_tags`: те трогают только справочник (цвет, закрепление),
    а сами задачи settings.py не видит — про хранилище знает только движок.

    Строки задач эта функция больше не переписывает: `tags`/`task_tags` в БД
    хранят тег через id, а не строкой на самой задаче, так что и переименование,
    и слияние — операции только над справочником `tags`, ни одной задачи не
    касаются. См. `store.Store.rename_tag_everywhere`.
    """
    return get_store().rename_tag_everywhere(old_name, new_name)


def main():
    p = argparse.ArgumentParser(description="Движок шагов Yungdrung")
    p.add_argument("--today", help="подменить сегодняшнюю дату (для проверок)")
    sub = p.add_subparsers(dest="cmd", required=True)

    sub.add_parser("next", help="что требует внимания").set_defaults(func=cmd_next)
    sub.add_parser("feed", help="лента «Что сегодня»").set_defaults(func=cmd_feed)
    sub.add_parser("backlog", help="всё просроченное — разбор завала").set_defaults(func=cmd_backlog)

    bb = sub.add_parser("backlog-bulk",
                        help="массовые действия из разбора завала, вкладка «Списком» (R20)")
    bb.add_argument("op", choices=["defer", "done", "fail"])
    bb.add_argument("items", help='JSON-список [{"task":"...","step":1}, ...] '
                                  'или "-" для чтения из stdin')
    bb.add_argument("--reason", help="одна причина на всю пачку; обязательна для defer/fail")
    bb.add_argument("--to", help="новая дата для defer, формат 2026-08-24 или "
                                 "2026-08-24 15:00 — уже разобранная, не «+3»")
    bb.set_defaults(func=cmd_backlog_bulk)
    sub.add_parser("list", help="все задачи").set_defaults(func=cmd_list)
    r = sub.add_parser("refresh", help="пересчитать сводку во всех задачах (перед сборкой)")
    r.add_argument("--force", action="store_true",
                   help="переписать все файлы, даже если сводка не изменилась — "
                        "нужно после смены схемы, чтобы привести вольт к новому виду")
    r.set_defaults(func=cmd_refresh)

    t = sub.add_parser("templates", help="список шаблонов с предпросмотром")
    t.add_argument("--start", help="от какой даты считать предпросмотр")
    t.set_defaults(func=cmd_templates)

    sr = sub.add_parser("set-recurrence", help="прикрепить или снять повторение у шаблона")
    sr.add_argument("name")
    sr.add_argument("--rule", help='JSON правила с anchor, например {"anchor":"2026-09-01","freq":"monthly","bymonthday":[5]}')
    sr.add_argument("--clear", action="store_true", help="снять повторение")
    sr.set_defaults(func=cmd_set_recurrence)

    st = sub.add_parser("save-template", help="создать шаблон с нуля или переписать существующий")
    st.add_argument("json", help='JSON шаблона или "-" для чтения из stdin, форма как у cmd_create')
    st.set_defaults(func=cmd_save_template)

    tp = sub.add_parser("template-preview", help="какие даты дадут шаги ещё не сохранённого шаблона")
    tp.add_argument("json", help='JSON шаблона или "-" для чтения из stdin')
    tp.add_argument("--start", help="дата старта для примера, по умолчанию сегодня")
    tp.set_defaults(func=cmd_template_preview)

    rp = sub.add_parser("parse-recurrence", help="разобрать «каждый вторник» в правило")
    rp.add_argument("text")
    rp.set_defaults(func=cmd_recurrence_parse)

    rc = sub.add_parser("recur", help="прогнать шаблоны с правилом повторения")
    rc.add_argument("--name", help="только один шаблон")
    rc.add_argument("--force", action="store_true",
                    help="создать очередной цикл, даже если предыдущий не закрыт "
                         "(только вместе с --name — кнопка про один конкретный цикл)")
    rc.add_argument("--limit", type=int, help="сколько решений максимум за прогон")
    rc.set_defaults(func=cmd_recur)

    ft = sub.add_parser("from-template", help="завести задачу из шаблона")
    ft.add_argument("name")
    ft.add_argument("--start", help="дата старта, по умолчанию сегодня")
    ft.add_argument("--title", help="имя задачи, по умолчанию имя шаблона")
    ft.set_defaults(func=cmd_from_template)

    tt = sub.add_parser("template-from-task", help="сделать шаблон из задачи")
    tt.add_argument("task")
    tt.add_argument("--name", help="имя шаблона, по умолчанию имя задачи")
    tt.set_defaults(func=cmd_template_from_task)

    c = sub.add_parser("create", help="завести задачу из JSON (её же зовёт форма)")
    c.add_argument("json", help='JSON или "-" для stdin')
    c.set_defaults(func=cmd_create)

    u = sub.add_parser("update", help="править задачу из JSON — карточка, не статусы шагов")
    u.add_argument("task")
    u.add_argument("json", help='JSON или "-" для stdin')
    u.add_argument("--force", action="store_true",
                   help="разрешить пропажу шага из данных без явного «снять»")
    u.set_defaults(func=cmd_update)

    cn = sub.add_parser("cancel", help="отменить задачу целиком")
    cn.add_argument("task")
    cn.add_argument("--reason")
    cn.set_defaults(func=cmd_cancel)

    dl = sub.add_parser("delete", help="удалить задачу насовсем")
    dl.add_argument("task")
    dl.set_defaults(func=cmd_delete)

    at = sub.add_parser("attach", help="прикрепить файл к задаче, шагу или шаблону")
    at.add_argument("task", nargs="?", help="название задачи; либо --template")
    at.add_argument("file", help="путь к файлу на диске")
    at.add_argument("--step", help="id шага, если не вся задача")
    at.add_argument("--template", help="название шаблона вместо задачи")
    at.add_argument("--caption", help="короткая подпись")
    at.set_defaults(func=cmd_attach)

    al = sub.add_parser("attachments", help="список вложений задачи, шага или шаблона")
    al.add_argument("task", nargs="?", help="название задачи; либо --template")
    al.add_argument("--step", help="id шага, если не вся задача")
    al.add_argument("--template", help="название шаблона вместо задачи")
    al.set_defaults(func=cmd_attachments)

    ad = sub.add_parser("attachment-delete", help="удалить вложение по id")
    ad.add_argument("id", type=int)
    ad.set_defaults(func=cmd_attachment_delete)

    ro = sub.add_parser("reopen", help="отменить закрытие шага (сделан → снова открыт)")
    ro.add_argument("task")
    ro.add_argument("step")
    ro.set_defaults(func=cmd_reopen)

    s = sub.add_parser("show", help="одна задача целиком")
    s.add_argument("task")
    s.set_defaults(func=cmd_show)

    d = sub.add_parser("done", help="шаг сделан")
    d.add_argument("task"); d.add_argument("step"); d.add_argument("--reason")
    d.set_defaults(func=cmd_done)

    n = sub.add_parser("notdone", help="шаг не сделан, остаётся открытым")
    n.add_argument("task"); n.add_argument("step")
    n.add_argument("--reason", help="почему: не дозвонился, ушёл в отпуск, было некогда")
    n.add_argument("--to", help="когда спросить снова; по умолчанию завтра")
    n.set_defaults(func=cmd_notdone)

    f = sub.add_parser("defer", help="перенести шаг на дату")
    f.add_argument("task"); f.add_argument("step")
    f.add_argument("--to", required=True); f.add_argument("--reason")
    f.set_defaults(func=cmd_defer)

    fl = sub.add_parser("fail", help="не будет сделано — шаг провален, задача идёт дальше")
    fl.add_argument("task")
    fl.add_argument("step")
    fl.add_argument("--reason", required=True, help="причина обязательна")
    fl.set_defaults(func=cmd_fail)

    k = sub.add_parser("skip", help="снять шаг")
    k.add_argument("task"); k.add_argument("step"); k.add_argument("--reason")
    k.set_defaults(func=cmd_skip)

    ar = sub.add_parser("archive", help="история: закрытые и отменённые задачи")
    ar.add_argument("--tag", help="только с этим тегом")
    ar.add_argument("--since", help="не раньше этой даты")
    ar.add_argument("--until", help="не позже этой даты")
    ar.set_defaults(func=cmd_archive)

    sr = sub.add_parser("search", help="поиск по задачам, заметкам и базе знаний")
    sr.add_argument("text")
    sr.add_argument("--kind", choices=["task", "kb_note"], help="только этот вид")
    sr.add_argument("--limit", type=int, default=50)
    sr.set_defaults(func=cmd_search)

    ri = sub.add_parser("reindex", help="собрать поисковый индекс заново")
    ri.set_defaults(func=cmd_reindex)

    mk = sub.add_parser("migrate-kb",
                        help="перенести базу знаний из База/*.md в таблицы (этап b)")
    mk.set_defaults(func=cmd_migrate_kb)

    ks = sub.add_parser("kb-scan", help="гипотезы упоминаний записей базы знаний в тексте")
    ks.add_argument("--text", default="")
    ks.add_argument("--source-type")
    ks.add_argument("--source-id")
    ks.set_defaults(func=cmd_kb_scan)

    kc = sub.add_parser("kb-confirm", help="подтвердить гипотезы, проставить ссылки в тексте")
    kc.add_argument("--source-type", required=True)
    kc.add_argument("--source-id", required=True)
    kc.add_argument("--mentions", required=True,
                    help='JSON-список гипотез (эхо того, что вернул kb-scan) или "-" для stdin')
    kc.set_defaults(func=cmd_kb_confirm)

    kr = sub.add_parser("kb-reject", help="отклонить гипотезу или заглушить слово целиком")
    kr.add_argument("--mention", required=True, help='JSON гипотезы или "-" для stdin')
    kr.add_argument("--mute", action="store_true", help="слово никогда не ссылка, у любой записи")
    kr.set_defaults(func=cmd_kb_reject)

    x = sub.add_parser("export", help="выгрузить весь вольт в Excel")
    x.add_argument("--to", help="куда писать; по умолчанию — «Выгрузка <дата>.xlsx» "
                                "в корне вольта")
    x.set_defaults(func=cmd_export)

    bk = sub.add_parser("backup", help="снять резервную копию вольта (R25)")
    bk.add_argument("--dest", help="куда класть копии; по умолчанию рядом с вольтом")
    bk.add_argument("--keep", type=int, help="сколько копий хранить, по умолчанию 7")
    bk.add_argument("--force", action="store_true",
                    help="снять копию сейчас, не глядя на расписание")
    bk.set_defaults(func=cmd_backup)

    bl = sub.add_parser("backups", help="список сделанных копий")
    bl.add_argument("--dest", help="папка копий; по умолчанию рядом с вольтом")
    bl.set_defaults(func=cmd_backup_list)

    rs = sub.add_parser("restore", help="восстановить вольт из копии — перезаписывает данные")
    rs.add_argument("file", help="путь к архиву копии")
    rs.set_defaults(func=cmd_backup_restore)

    ej = sub.add_parser("export-json", help="выгрузить всю базу в JSON (R11)")
    ej.add_argument("--to", help="куда писать файл выгрузки; по умолчанию в корне вольта")
    ej.set_defaults(func=cmd_export_json)

    args = p.parse_args()
    today = date.fromisoformat(args.today) if args.today else date.today()
    print(json.dumps(args.func(args, today), ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
